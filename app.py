"""
app.py v6 — Predimensionamiento de Cimentaciones Superficiales
Fase 1: UX mejorada, flujo controlado, vigas editables

PASO 0: Cargar modelo → mostrar columnas en planta
PASO 1: Usuario clasifica → botón Aplicar clasificación
PASO 2: Vigas de enlace editables → botón Aplicar vigas
PASO 3: Ejecutar diseño → resultados separados
"""
import streamlit as st
import plotly.graph_objects as go
import pandas as pd
import json, math, tempfile, os, io
import base64
from engine import (
    read_model,
    resolve_basis_selection,
    run_design,
    normalize_reactions_df,
)
from isolated import propose_rebar, infer_column_axis
from tie_system import deduce_tie_beams
from parser import auto_classify_patterns, classify_from_user
from export_s2k import export_foundation_s2k
# from export_e2k import export_foundation_e2k  # temporalmente suspendido

st.set_page_config(
    page_title="Cimentaciones",
    page_icon="assets/logo2.png",
    layout="wide"
)

# ════════════════════════════════════════════════
# HELPERS
# ════════════════════════════════════════════════
def apply_theme(fig, height=500, equal_axes=True):
    """Tema visual claro para todas las gráficas."""
    xa = dict(title_font=dict(size=11), gridcolor='rgba(0,0,0,0)',
              showgrid=False, zeroline=False, linecolor='#d1d5db', linewidth=1,
              tickfont=dict(size=9, color='#374151'))
    ya = dict(title_font=dict(size=11), gridcolor='rgba(0,0,0,0)',
              showgrid=False, zeroline=False, linecolor='#d1d5db', linewidth=1,
              tickfont=dict(size=9, color='#374151'))
    if equal_axes:
        xa['scaleanchor'] = 'y'; xa['scaleratio'] = 1
    fig.update_layout(
        plot_bgcolor='white', paper_bgcolor='white',
        font=dict(color='#1f2937', size=11),
        xaxis=xa, yaxis=ya, height=height,
        margin=dict(l=50, r=30, t=40, b=40),
    )
    return fig


def get_preliminary_position(col, cl_info):
    """Posición del centro de zapata conceptual según clasificación."""
    cx, cy = col['x'], col['y']
    bx, by = col['bx'], col['by']
    dim = round(2 * max(bx, by), 2)
    margin = 0.05
    loc = cl_info.get('location', 'concentrica')
    side = cl_info.get('side', '')
    corner = cl_info.get('corner', '')
    zx, zy = cx, cy
    if loc == 'medianera':
        if side == 'X+':   zx = cx + bx/2 + margin - dim/2
        elif side == 'X-': zx = cx - bx/2 - margin + dim/2
        elif side == 'Y+': zy = cy + by/2 + margin - dim/2
        elif side == 'Y-': zy = cy - by/2 - margin + dim/2
    elif loc == 'esquinera':
        if 'X+' in corner:   zx = cx + bx/2 + margin - dim/2
        elif 'X-' in corner: zx = cx - bx/2 - margin + dim/2
        if 'Y+' in corner:   zy = cy + by/2 + margin - dim/2
        elif 'Y-' in corner: zy = cy - by/2 - margin + dim/2
    return round(zx, 3), round(zy, 3), dim


LOC_COLORS = {
    'concentrica': ('#6b7280', 'rgba(107,114,128,0.15)'),
    'medianera':   ('#3b82f6', 'rgba(59,130,246,0.15)'),
    'esquinera':   ('#f97316', 'rgba(249,115,22,0.15)'),
}

STATUS_COLORS = {
    'PRELIMINAR_OK':              ('#059669', 'rgba(5,150,105,0.10)'),
    'PRELIMINAR_OK_COMBINADA':    ('#059669', 'rgba(5,150,105,0.10)'),
    'REVISION_EXCENTRICIDAD':     ('#2563eb', 'rgba(37,99,235,0.10)'),
    'REVISION_COMBINADA':         ('#7c3aed', 'rgba(124,58,237,0.10)'),
    'REVISAR_h':                  ('#d97706', 'rgba(217,119,6,0.10)'),
    'NO_CUMPLE':                  ('#dc2626', 'rgba(220,38,38,0.10)'),
}


def ties_dict_to_table(ties, columns):
    """Convierte el dict de ties del engine a filas de tabla editable."""
    rows = []
    for col in columns:
        jid = col['joint']
        t = ties.get(jid, {})
        if not t.get('needs_tie', False):
            continue
        cx, cy = 'Ninguno', 'Ninguno'
        if t.get('is_corner'):
            tx = t.get('tie_x', {})
            ty = t.get('tie_y', {})
            if tx.get('tie_to'): cx = f"J{tx['tie_to']}"
            if ty.get('tie_to'): cy = f"J{ty['tie_to']}"
        else:
            if t.get('tie_to'):
                if t.get('tie_dir') == 'X':
                    cx = f"J{t['tie_to']}"
                else:
                    cy = f"J{t['tie_to']}"
        rows.append({'Nodo': f'J{jid}', 'Conecta_X': cx, 'Conecta_Y': cy, 'Origen': 'auto'})
    return rows


def table_to_ties_dict(tie_rows, columns):
    """Convierte la tabla editable del usuario a dict de ties para el engine."""
    col_map = {c['joint']: c for c in columns}
    # Build ID lookup: J5 → 5
    def clean(val):
        s = str(val).strip()
        if s == 'Ninguno':
            return s
        if s.startswith('J') and len(s) > 1 and s[1:].isdigit():
            return s[1:]
        return s

    ties = {}
    for row in tie_rows:
        jid = clean(row['Nodo'])
        cx = clean(row.get('Conecta_X', 'Ninguno'))
        cy = clean(row.get('Conecta_Y', 'Ninguno'))
        col = col_map.get(jid)
        if not col:
            continue
        has_x = cx != 'Ninguno' and cx in col_map
        has_y = cy != 'Ninguno' and cy in col_map
        if not has_x and not has_y:
            continue
        if has_x and has_y:
            tx = col_map[cx]; ty = col_map[cy]
            ties[jid] = {
                'needs_tie': True, 'is_corner': True,
                'scheme_suggested': 'doble_viga',
                'tie_x': {'tie_to': cx, 'tie_dir': 'X',
                          'tie_dist': round(abs(tx['x'] - col['x']), 2),
                          'tie_x': tx['x'], 'tie_y': tx['y']},
                'tie_y': {'tie_to': cy, 'tie_dir': 'Y',
                          'tie_dist': round(abs(ty['y'] - col['y']), 2),
                          'tie_x': ty['x'], 'tie_y': ty['y']},
            }
        elif has_x:
            tx = col_map[cx]
            ties[jid] = {
                'needs_tie': True, 'tie_to': cx, 'tie_dir': 'X',
                'tie_dist': round(abs(tx['x'] - col['x']), 2),
                'tie_x': tx['x'], 'tie_y': tx['y'],
                'scheme_suggested': 'viga_X',
            }
        else:
            ty = col_map[cy]
            ties[jid] = {
                'needs_tie': True, 'tie_to': cy, 'tie_dir': 'Y',
                'tie_dist': round(abs(ty['y'] - col['y']), 2),
                'tie_x': ty['x'], 'tie_y': ty['y'],
                'scheme_suggested': 'viga_Y',
            }
    # Nodos sin vigas
    for c in columns:
        if c['joint'] not in ties:
            ties[c['joint']] = {'needs_tie': False}
    return ties


# ════════════════════════════════════════════════
# SIDEBAR
# ════════════════════════════════════════════════
with st.sidebar:
    # 🔹 Logo con altura controlada
    st.image("assets/logo.png", use_container_width=True)  # ajusta el ancho aquí

    # 🔹 Texto pegado al logo
    st.markdown(
        """
        <div style='font-size:11px; color:#6b7280; text-align:center; margin-top:-100px; line-height:1.2;'>
        Aplicación desarrollada con fines didácticos por<br>
        SmartCouplers MG SAS.<br>
        No está auditada ni validada.<br>
        NO USAR PARA FINES PROFESIONALES!<br>
        Ultima Modificacion 07-07042026<br>
        </div>
        """,
        unsafe_allow_html=True
    )
    st.title("⚙️ Parámetros")
    R = st.number_input("R", value=7.0, step=0.5)
    fc = st.number_input("f'c (MPa)", value=21.0, step=1.0)
    fy = st.number_input("fy (MPa)", value=420.0, step=10.0)
    rec = st.number_input("Rec. (cm)", value=7.5, step=0.5)
    Df = st.number_input("Df (m)", value=1.5, step=0.1)
    gs = st.number_input("γ suelo", value=18.0)
    gc = st.number_input("γ concreto", value=24.0)
    qadm_1 = st.number_input("qadm D+L (kPa)", value=180.0, step=5.0)
    qadm_2 = st.number_input("qadm D+L máx", value=215.0, step=5.0)
    qadm_3 = st.number_input("qadm D+L+E", value=358.0, step=5.0)
    h_min = st.number_input("h mín (m)", value=0.30, step=0.05)
    dim_min = st.number_input("Dim mín B/L (m)", value=0.60, step=0.05)
    st.divider()
    st.subheader("Estabilidad")
    mu = st.number_input("μ fricción suelo", value=0.40, step=0.05, min_value=0.10, max_value=0.80)
    fs_volc_est = st.number_input("FS volc. estático", value=2.0, step=0.1)
    fs_volc_sis = st.number_input("FS volc. sísmico", value=1.5, step=0.1)
    fs_desl_est = st.number_input("FS desl. estático", value=1.5, step=0.1)
    fs_desl_sis = st.number_input("FS desl. sísmico", value=1.2, step=0.1)
    params = {'R':R, 'fc':fc*1000, 'fy':fy*1000, 'rec':rec/100, 'gamma_c':gc,
              'gamma_s':gs, 'Df':Df, 'h_min':h_min, 'dim_min':dim_min,
              'qadm_1':qadm_1, 'qadm_2':qadm_2, 'qadm_3':qadm_3,
              'mu': mu,
              'fs_volc_min': {'q1': fs_volc_est, 'q2': fs_volc_est, 'q3': fs_volc_sis},
              'fs_desl_min': {'q1': fs_desl_est, 'q2': fs_desl_est, 'q3': fs_desl_sis}}
    st.divider()
    st.subheader("📤 Exportación del modelo")
    sap_units = st.selectbox("Unidades .$2k (SAP2000)", ["KN, m, C"], index=0)
    pedestal_h_exp = st.number_input("Altura pedestal exportación (m)", value=0.50, step=0.05)
    k_subgrade_exp = st.number_input("k balasto exportación (kN/m³)", value=12000.0, step=500.0)
    alpha_xy_exp = st.number_input("α resortes horizontales", value=0.35, step=0.05, min_value=0.0)
    z_top_exp = st.number_input("Cota superior del pedestal Z (m)", value=0.0, step=0.10)
    tie_b_exp = st.number_input("b viga enlace respaldo (m)", value=0.30, step=0.05)
    tie_h_exp = st.number_input("h viga enlace respaldo (m)", value=0.50, step=0.05)
    st.caption("El exportador usará un único material de concreto generado a partir de f'c.")


# ════════════════════════════════════════════════
# MAIN — CARGA DE ARCHIVO
# ════════════════════════════════════════════════

st.markdown(
    """
    <div style='text-align: center;'>
        <img src="data:image/png;base64,{}" width="180">
        <h1 style='margin-top: -30px;'>Diseño de cimentaciones superficiales</h1>
    </div>
    """.format(
        base64.b64encode(open("assets/logo2.png", "rb").read()).decode()
    ),
    unsafe_allow_html=True
)

uploaded = st.file_uploader(
    "📁 Subir modelo estructural (.$2k de SAP2000 o .e2k / .$et de ETABS)",
    type=None,
)
if uploaded is None:
    st.info("Sube un archivo .$2k de SAP2000 o .e2k / .$et de ETABS.")
    st.stop()

raw = uploaded.read(); uploaded.seek(0)
_is_etabs = b'PROGRAM "ETABS"' in raw[:3000] or b'PROGRAM  "ETABS"' in raw[:3000]
_is_sap   = b'TABLE:' in raw[:1500]
if not _is_etabs and not _is_sap:
    st.error("Archivo no reconocido. Se esperan formatos SAP2000 (.$2k) o ETABS (.e2k / .$et).")
    st.stop()

_suffix = '.e2k' if _is_etabs else '.s2k'
# ── PASO 0: Leer modelo + diagnosticar fuente ──
with tempfile.NamedTemporaryFile(delete=False, suffix=_suffix) as tmp:
    tmp.write(raw)
    tmp_path = tmp.name

try:
    file_changed = st.session_state.get('_file_name') != uploaded.name
    if 'model_data' not in st.session_state or file_changed:
        md = read_model(tmp_path, params)
        st.session_state['model_data'] = md
        st.session_state['_file_name'] = uploaded.name

        # Reiniciar selección de base cuando cambia archivo
        st.session_state.pop('basis_selected', None)
        st.session_state.pop('basis_mode_user', None)

        # Inicializar clasificaciones por defecto con las columnas actuales
        cl_default = {}
        for c in md.get('design_entities', []):
            jid = str(c['joint'])
            cl_default[jid] = {
                'location': 'concentrica',
                'side': '',
                'corner': ''
            }
        st.session_state['classifications'] = cl_default

        # Inicializar clasificación de patrones de carga (auto-detección)
        _lpats_raw = md.get('_lpats_raw', md.get('_lp', {}))
        _lp_auto = auto_classify_patterns(_lpats_raw)
        st.session_state['lp_user_assignment'] = _lp_auto
        st.session_state.pop('lp_class_confirmed', None)
        # Limpiar estado downstream
        for _k in ['class_applied', 'tie_beams_table', 'ties_applied', 'results', 'results_verify', 'dims_design']:
            st.session_state.pop(_k, None)

    md = st.session_state['model_data']
finally:
    os.unlink(tmp_path)


def _lp_class_section(source_dict, md, params, label="cargas del modelo"):
    """Renders load classification UI; stops script until confirmed.
    source_dict: {name: type_str} — use {case: ""} for reactions output cases.
    """
    from collections import Counter as _Counter
    _lp_categories = ['D', 'SD', 'L', 'Lr', 'Le', 'Wx+', 'Wx-', 'Wy+', 'Wy-', 'Sx', 'Sy', 'Ignorar']

    if not source_dict:
        return

    # Detect source change → reset state
    _src_keys = tuple(sorted(source_dict.keys()))
    if st.session_state.get('_lp_source_keys') != _src_keys:
        st.session_state['_lp_source_keys'] = _src_keys
        st.session_state['lp_user_assignment'] = auto_classify_patterns(source_dict)
        st.session_state.pop('lp_class_confirmed', None)

    lp_assign = st.session_state.get('lp_user_assignment', auto_classify_patterns(source_dict))

    st.subheader(f"📂 Clasificación de {label}")
    lp_rows = []
    for name, typ in source_dict.items():
        lp_rows.append({
            'Nombre': name,
            'Tipo (archivo)': typ if typ else '—',
            'Categoría diseño': lp_assign.get(name, 'Ignorar'),
        })

    with st.form('lp_class_form', clear_on_submit=False):
        lp_edited = st.data_editor(
            pd.DataFrame(lp_rows),
            column_config={
                'Nombre':           st.column_config.TextColumn(disabled=True, width='medium'),
                'Tipo (archivo)':   st.column_config.TextColumn(disabled=True, width='medium'),
                'Categoría diseño': st.column_config.SelectboxColumn(
                    options=_lp_categories, required=True, width='medium'),
            },
            use_container_width=True, hide_index=True, key='lp_class_editor',
        )
        lp_confirm = st.form_submit_button("✅ Confirmar clasificación de cargas",
                                           type='primary', use_container_width=True)

    if lp_confirm:
        new_assign = {row['Nombre']: row['Categoría diseño'] for _, row in lp_edited.iterrows()}
        warns = []
        cnt = _Counter(new_assign.values())
        for cat in ('Sx', 'Sy'):
            if cnt.get(cat, 0) > 1:
                warns.append(f"⚠️ Múltiples cargas asignadas a '{cat}': "
                             f"{', '.join(p for p, c in new_assign.items() if c == cat)}")
        if 'D' not in new_assign.values():
            warns.append("⚠️ Ninguna carga asignada como Muerta (D).")
        for w in warns:
            st.warning(w)

        # ── Build rename map: original → canonical (None = ignorar) ───────────
        from export_utils import CAT_TO_DESTYPE as _CAT_DT
        _cat_to_cl_key = {
            'D': 'dead', 'SD': 'superdead', 'L': 'live', 'Lr': 'live_roof', 'Le': 'live_eq',
            'Sx': 'seismic_x', 'Sy': 'seismic_y',
            'Wx+': 'wind_xp', 'Wx-': 'wind_xn', 'Wy+': 'wind_yp', 'Wy-': 'wind_yn',
        }
        _cat_ctr: dict = {}
        _rmap: dict = {}   # original → canonical_name | None
        for _orig, _cat in new_assign.items():
            if _cat == 'Ignorar':
                _rmap[_orig] = None
                continue
            _cat_ctr[_cat] = _cat_ctr.get(_cat, 0) + 1
            _n = _cat_ctr[_cat]
            _rmap[_orig] = _cat if _n == 1 else f"{_cat}_{_n}"

        # ── Guardar originales (solo la primera vez) ───────────────────────────
        if '_jloads_orig' not in md:
            md['_jloads_orig'] = {j: dict(v) for j, v in (md.get('_jloads') or {}).items()}
        if '_lp_orig' not in md:
            md['_lp_orig'] = dict(md.get('_lp') or {})

        # ── Renombrar _jloads desde los originales ─────────────────────────────
        _new_jloads: dict = {}
        for _jid, _pj in md['_jloads_orig'].items():
            _renamed: dict = {}
            for _orig, _fv in _pj.items():
                _exp = _rmap.get(_orig, _orig)   # casos desconocidos: mantener nombre
                if _exp is None:
                    continue                      # ignorados
                if _exp in _renamed:
                    for _k in _renamed[_exp]:
                        _renamed[_exp][_k] = _renamed[_exp].get(_k, 0) + _fv.get(_k, 0)
                else:
                    _renamed[_exp] = dict(_fv)
            if _renamed:
                _new_jloads[_jid] = _renamed

        # ── Renombrar _lp desde los originales ────────────────────────────────
        _new_lp: dict = {}
        for _orig, _dt in md['_lp_orig'].items():
            _exp = _rmap.get(_orig)
            if _exp is None:
                continue                          # ignorados
            if _exp not in _new_lp:
                _new_lp[_exp] = _CAT_DT.get(new_assign.get(_orig, ''), _dt)
        # Conservar patrones desconocidos (no en new_assign)
        for _orig, _dt in md['_lp_orig'].items():
            if _orig not in _rmap and _orig not in _new_lp:
                _new_lp[_orig] = _dt

        # ── Construir _cl con nombres canónicos ───────────────────────────────
        _new_cl: dict = {_k: [] for _k in list(_cat_to_cl_key.values()) + ['other']}
        for _orig, _cat in new_assign.items():
            if _cat == 'Ignorar':
                continue
            _exp = _rmap.get(_orig)
            if _exp is None:
                continue
            _cl_key = _cat_to_cl_key.get(_cat, 'other')
            if _exp not in _new_cl[_cl_key]:
                _new_cl[_cl_key].append(_exp)

        # ── assignment canónico para export (canonical → category) ────────────
        _canonical_assign = {
            _exp: _cat
            for _orig, _cat in new_assign.items()
            if (_exp := _rmap.get(_orig)) is not None and _cat != 'Ignorar'
        }

        md['_jloads']            = _new_jloads
        md['_lp']                = _new_lp
        md['_cl']                = _new_cl
        md['_lp_user_assignment'] = _canonical_assign   # {canonical: category} para export
        md['_rename_map']         = _rmap               # {original: canonical | None} para run_design

        # ── Renombrar design_jloads si ya está en md (modo joint_loads) ───────
        from export_utils import apply_rename_to_jloads as _arj
        if md.get('design_jloads'):
            md['design_jloads'] = _arj(md['design_jloads'], _rmap)

        new_combos = __import__('parser').gen_combos(_new_cl, R=params['R'])
        md['n_combos'] = {'ADS': len(new_combos['ADS']), 'LRFD': len(new_combos['LRFD'])}

        st.session_state['lp_user_assignment'] = new_assign   # {original: category} para UI
        st.session_state['lp_class_confirmed'] = True
        st.session_state['model_data'] = md
        st.rerun()

    if not st.session_state.get('lp_class_confirmed', False):
        st.info("👆 Revise y confirme la clasificación de cargas para continuar.")
        st.stop()

    # Mostrar resumen con nombres CANÓNICOS (no los del archivo de origen)
    canonical_assign = md.get('_lp_user_assignment', {})   # {canonical: category}
    if canonical_assign:
        by_cat: dict = {}
        for _exp, _cat in canonical_assign.items():
            by_cat.setdefault(_cat, []).append(_exp)
        st.success("✅ Clasificación confirmada — " +
                   ' | '.join(f"**{c}**: {', '.join(sorted(ps))}"
                               for c, ps in sorted(by_cat.items())))

# ── Mostrar advertencias generales del modelo ──
if md.get('warnings'):
    for w in md['warnings']:
        st.warning(w)

# ── Diagnóstico Fase 0 ──
basis_mode = md.get('basis_mode', 'invalid_model')
status_message = md.get('status_message', '')
n_j = md.get('joint_load_count', 0)
n_r = md.get('restraint_count', 0)
n_wj = md.get('_diag', {}).get('wall_joint_count', 0)
n_wr = md.get('_diag', {}).get('wall_restraint_count', 0)

st.subheader("🧭 Paso 0 — Diagnóstico del modelo")
if status_message:
    if basis_mode == 'invalid_model':
        st.error(status_message)
    elif basis_mode == 'ask_user':
        st.warning(status_message)
    else:
        st.info(status_message)

c1, c2, c3, c4, c5, c6 = st.columns(6)
c1.metric("Candidatos por cargas nodales", n_j)
c2.metric("Candidatos por restricciones", n_r)
c3.metric("Muros por cargas", n_wj)
c4.metric("Muros por restricciones", n_wr)
c5.metric("Modo detectado", basis_mode)
c6.metric("Entidades puntuales", md.get("n_foundation_entities", len(md.get("columns", []))))

# ── Resolver caso ask_user ──
if basis_mode == 'ask_user':
    default_ix = 0
    current_choice = st.radio(
        "El modelo tiene ambas fuentes posibles. Seleccione la base de trabajo:",
        options=['joint_loads', 'support_reactions'],
        index=default_ix,
        format_func=lambda x: (
            "A — Fuerzas en nudos" if x == 'joint_loads'
            else "B — Restricciones/apoyos base"
        ),
        horizontal=True,
        key='basis_mode_user'
    )

    if st.button("✅ Aplicar fuente seleccionada", type="primary", use_container_width=True):
        md = resolve_basis_selection(st.session_state['model_data'], current_choice, params)
        st.session_state['model_data'] = md
        st.session_state['basis_selected'] = True

        # Reinicializar clasificaciones según nueva base
        cl_default = {}
        for c in md.get('design_entities', []):
            jid = str(c['joint'])
            cl_default[jid] = {
                'location': 'concentrica',
                'side': '',
                'corner': ''
            }
        st.session_state['classifications'] = cl_default

        # Limpiar estado downstream
        for k in ['class_applied', 'tie_beams_table', 'ties_applied', 'results', 'results_verify', 'dims_design']:
            st.session_state.pop(k, None)

        st.rerun()

    if not st.session_state.get('basis_selected', False):
        st.info("Seleccione la fuente de trabajo para continuar.")
        st.stop()

    md = st.session_state['model_data']
    basis_mode = md.get('basis_mode', current_choice)

# ── Modelo inválido: detener ──
if basis_mode == 'invalid_model':
    st.error("El proceso se detiene porque el modelo no es representativo para cimentaciones.")
    st.stop()

columns = md.get('columns', [])

# IMPORTANTÍSIMO:
# usar la base original de muros activos, no la versión ya segmentada/persistida
wall_entities_base = list((md.get("_active_wall_entities") or {}).values())
wall_entities = wall_entities_base

wall_resultants = md.get("wall_resultants", {})
wall_foundation_entities = md.get("wall_foundation_entities", [])
design_entities = md.get("design_entities", [])
wall_segment_audit = md.get("wall_segment_audit", [])

md["wall_entities"] = wall_entities
md["wall_resultants"] = wall_resultants
md["wall_foundation_entities"] = wall_foundation_entities


# ── Mensajes por modo ──
if basis_mode == 'joint_loads':
    st.success(
        f"✅ Fuente activa: fuerzas en nudos | "
        f"{len(design_entities)} entidades de diseño detectadas | "
        f"{md['n_combos']['ADS']}+{md['n_combos']['LRFD']} combinaciones"
    )
    # Clasificación de cargas usando patrones del modelo
    _lpats_raw = md.get('_lpats_raw', md.get('_lp', {}))
    _lp_class_section(_lpats_raw, md, params, "patrones de carga del modelo")
elif basis_mode == 'support_reactions':
    st.success(
        f"✅ Fuente activa: restricciones/apoyos base | "
        f"{len(design_entities)} entidades de diseño detectadas"
    )

    _file_fmt = md.get('file_format', 'sap2000')
    _react_label = "SAP2000" if _file_fmt == 'sap2000' else "ETABS"
    file_xlsx = st.file_uploader(
        f"📥 Suba el Excel de reacciones exportado desde {_react_label}",
        type=["xlsx"],
        key="reactions_xlsx"
    )

    if file_xlsx is None:
        st.warning("Debe cargar el Excel de reacciones para continuar.")
        st.stop()

    file_xlsx.seek(0)
    try:
        df_react_preview = pd.read_excel(file_xlsx, sheet_name="Joint Reactions", header=1)
    except Exception:
        file_xlsx.seek(0)
        df_react_preview = pd.read_excel(file_xlsx, header=1)

    # Normalizar nombres de columna (ETABS usa nombres diferentes a SAP2000)
    df_react_preview = normalize_reactions_df(df_react_preview)

    # Para ETABS: filtrar solo piso base si hay columna 'Story'
    if _file_fmt == 'etabs' and 'Story' in df_react_preview.columns:
        _base_story = md.get('_etabs_base_story', 'N+0.0')
        df_react_preview = df_react_preview[
            df_react_preview['Story'].astype(str).str.strip() == _base_story
        ].copy()
        if df_react_preview.empty:
            st.warning(f"No se encontraron reacciones para el piso base '{_base_story}'. Se usarán todos los datos.")
            file_xlsx.seek(0)
            try:
                df_react_preview = normalize_reactions_df(
                    pd.read_excel(file_xlsx, sheet_name="Joint Reactions", header=1)
                )
            except Exception:
                pass

    st.session_state["reactions_file"] = file_xlsx
    st.session_state["reactions_df"] = df_react_preview

    st.info("Excel de reacciones cargado correctamente.")

    from engine import (
        build_jloads_from_sap_reactions_excel,
        segment_wall_entities_if_needed,
        build_wall_resultants,
        build_wall_foundation_entities,
        build_wall_as_column_entities,
        build_wall_as_column_jloads,
        merge_design_entities,
        merge_design_jloads,
    )

    # 1) Reacciones completas para TODAS las entidades puntuales reales
    base_jloads_full = build_jloads_from_sap_reactions_excel(
        df_react_preview,
        md.get("foundation_entities", []),
    )

    # B) Reacciones específicas de muros activos
    wall_foundation_entities_tmp = []
    used_wall_joints = set()

    for w in wall_entities_base:
        for jid in w.get("active_base_joints", []) or w.get("base_joints", []):
            jid_s = str(jid)
            if jid_s in used_wall_joints:
                continue
            wall_foundation_entities_tmp.append({
                "joint": jid_s,
                "entity_type": "point",
            })
            used_wall_joints.add(jid_s)

    wall_jloads_react = build_jloads_from_sap_reactions_excel(
        df_react_preview,
        wall_foundation_entities_tmp
    )

    # 2) Segmentar muros compuestos usando esas reacciones
    wall_entities_segmented, wall_segment_audit = segment_wall_entities_if_needed(
        wall_entities_base,
        md["_shell_entities"],
        md["_joints"],
        wall_jloads_react,
    )

    wall_entities = wall_entities_segmented
    md["wall_segment_audit"] = wall_segment_audit

    # 3) Resultantes por tramo
    wall_resultants = build_wall_resultants(
        wall_entities,
        wall_jloads_react,
        md["_joints"],
    )

    # 4) Entidades lineales
    wall_foundation_entities = build_wall_foundation_entities(
        wall_entities,
        wall_resultants,
        md["_joints"],
        default_support_width=params.get("dim_min", 0.60),
    )

    # 5) Muros equivalentes tipo columna
    wall_as_column_entities = build_wall_as_column_entities(
        wall_entities,
        wall_resultants,
        md["_joints"],
    )

    wall_as_column_jloads = build_wall_as_column_jloads(
        wall_as_column_entities
    )

    # 6) Entidades unificadas
    design_entities = merge_design_entities(
        columns,
        wall_as_column_entities,
    )

    design_jloads = merge_design_jloads(
        columns,
        wall_as_column_entities,
        base_jloads_full,
        wall_as_column_jloads,
    )

    # 7) Aplicar rename canónico a design_jloads si la clasificación ya fue confirmada
    _rmap_active = md.get("_rename_map", {})
    if _rmap_active:
        from export_utils import apply_rename_to_jloads as _arj2
        design_jloads = _arj2(design_jloads, _rmap_active)

    # 8) Persistir TODO en md para que sobreviva al rerun
    md["wall_entities"] = wall_entities
    md["wall_resultants"] = wall_resultants
    md["wall_foundation_entities"] = wall_foundation_entities
    md["wall_segment_audit"] = wall_segment_audit
    md["design_entities"] = design_entities
    md["design_jloads"] = design_jloads
    md["reactions_df"] = df_react_preview

    st.session_state["model_data"] = md

    # 8) Sincronizar clasificaciones con las entidades actuales
    current_cl = st.session_state.get("classifications", {})
    synced_cl = {}

    for ent in design_entities:
        jid = str(ent["joint"])
        prev = current_cl.get(jid, {})
        synced_cl[jid] = {
            "location": prev.get("location", "concentrica"),
            "side": prev.get("side", ""),
            "corner": prev.get("corner", ""),
            "ecc_x": prev.get("ecc_x", 0.0),
            "ecc_y": prev.get("ecc_y", 0.0),
            "ecc_dir": prev.get("ecc_dir", "ambas"),
            "mpx": prev.get("mpx", 0.0),
            "mpy": prev.get("mpy", 0.0),
            "vux": prev.get("vux", 0.0),
            "vuy": prev.get("vuy", 0.0),
        }

    st.session_state["classifications"] = synced_cl

    # ── Limpiar downstream SOLO si cambió realmente la población de entidades ──
    current_entity_ids = tuple(sorted(str(e["joint"]) for e in design_entities))
    prev_entity_ids = st.session_state.get("_design_entity_ids_support_reactions")

    if prev_entity_ids != current_entity_ids:
        for k in ["tie_beams_table", "ties_applied", "results", "results_verify", "dims_design"]:
            st.session_state.pop(k, None)

    st.session_state["_design_entity_ids_support_reactions"] = current_entity_ids

    # Clasificación usando los casos de carga del Excel (OutputCase)
    if 'OutputCase' in df_react_preview.columns:
        _react_cases = {c: "" for c in sorted(
            df_react_preview['OutputCase'].dropna().astype(str).unique())}
    else:
        _react_cases = md.get('_lpats_raw', md.get('_lp', {}))
    _lp_class_section(_react_cases, md, params, "casos de carga (Excel de reacciones)")

else:
    st.info(
        f"{len(columns)} candidatos detectados | "
        f"{md['n_combos']['ADS']}+{md['n_combos']['LRFD']} combinaciones"
    )

st.subheader("🧩 Entidades unificadas de diseño")
if design_entities:
    de_rows = []
    for e in design_entities:
        de_rows.append({
            "ID": e.get("id", ""),
            "Joint": e.get("joint", ""),
            "Familia": e.get("design_family", ""),
            "Tipo": e.get("type", ""),
            "X": round(float(e.get("x", 0.0)), 3),
            "Y": round(float(e.get("y", 0.0)), 3),
            "bx": round(float(e.get("bx", 0.0)), 3),
            "by": round(float(e.get("by", 0.0)), 3),
            "Fuente": e.get("source", ""),
        })
    st.dataframe(pd.DataFrame(de_rows), use_container_width=True)
else:
    st.info("No hay entidades unificadas de diseño para la base activa.")

# Lista de IDs para selectboxes
node_options = ['Ninguno'] + [f"J{c['joint']}" for c in design_entities]

if not design_entities:
    st.error("No hay entidades de diseño disponibles para clasificar.")
    st.stop()

# ════════════════════════════════════════════════
# PASO 1: CLASIFICAR
# ════════════════════════════════════════════════
st.header("📋 Paso 1 — Clasificar posición de cada zapata")

# Gráfico inicial: solo columnas
fig0 = go.Figure()
for c in design_entities:
    bx, by = c['bx'], c['by']
    fig0.add_shape(type="rect",
        x0=c['x']-bx/2, y0=c['y']-by/2, x1=c['x']+bx/2, y1=c['y']+by/2,
        line=dict(color='#374151', width=1.5), fillcolor='#d1d5db')
    fig0.add_annotation(x=c['x'], y=c['y']+by/2+0.15,
        text=f"<b>J{c['joint']}</b><br>{bx*100:.0f}×{by*100:.0f}",
        showarrow=False, font=dict(size=8, color='#374151'))
fig0.update_layout(title="Entidades de diseño del modelo (posición y sección)")
apply_theme(fig0, height=450)
st.plotly_chart(fig0, use_container_width=True)

# Tabla de clasificación editable
st.subheader("Clasificación por entidad")
st.caption("Modifique la localización de cada columna. "
           "Mpx/Mpy y Vux/Vuy muestran el máximo sísmico de referencia — ajuste si necesita.")

cl = st.session_state['classifications']
loc_options = ['concentrica', 'medianera', 'esquinera']
side_options = ['', 'X+', 'X-', 'Y+', 'Y-']
corner_options = ['', 'X+Y+', 'X+Y-', 'X-Y+', 'X-Y-']

# ── Calcular Mp/Vu de referencia = fuerzas sísmicas efectivas (Sx/R, Sy/R) ───
def _seismic_ref(jid_str, jloads_dict, cl_dict, R_factor):
    """
    Devuelve (mpx, mpy, vux, vuy) como los valores máximos resultantes de
    aplicar los patrones sísmicos con factor 1/R — exactamente como hace
    gen_combos con Ex(f) = fac("seismic_x", f/R).

    cl_dict: md['_cl']  →  {'seismic_x': ['Sx', ...], 'seismic_y': ['Sy', ...], ...}
    """
    mpx = mpy = vux = vuy = 0.0
    R = R_factor if R_factor and R_factor > 0 else 1.0

    # Patrones clasificados como sísmicos en X e Y
    sx_pats = cl_dict.get('seismic_x', [])
    sy_pats = cl_dict.get('seismic_y', [])

    joint_loads = jloads_dict.get(jid_str, {})

    for _pat in sx_pats:
        _fv = joint_loads.get(_pat, {})
        if not _fv:
            continue
        # Ex = Sx / R  →  My = M2/R,  Vx = F1/R
        mpx = max(mpx, abs(_fv.get('M2', 0)) / R)
        vux = max(vux, abs(_fv.get('F1', 0)) / R)

    for _pat in sy_pats:
        _fv = joint_loads.get(_pat, {})
        if not _fv:
            continue
        # Ey = Sy / R  →  Mx = M1/R,  Vy = F2/R
        mpy = max(mpy, abs(_fv.get('M1', 0)) / R)
        vuy = max(vuy, abs(_fv.get('F2', 0)) / R)

    return round(mpx, 2), round(mpy, 2), round(vux, 2), round(vuy, 2)

_jloads_ref  = md.get('_jloads') or md.get('design_jloads') or {}
_cl_ref      = md.get('_cl', {})
_R_ref       = params.get('R', 1.0)

edit_rows = []
for c in design_entities:
    jid = str(c['joint'])
    ci = cl.get(jid, {})
    # Usar valor guardado si el usuario ya lo definió (≠ 0), sino sugerir el sísmico
    _ref_mpx, _ref_mpy, _ref_vux, _ref_vuy = _seismic_ref(jid, _jloads_ref, _cl_ref, _R_ref)
    edit_rows.append({
        'Nodo': f'J{jid}',
        'Familia': c.get('design_family', 'column'),
        'Geometría': f"{c['bx']*100:.0f}×{c['by']*100:.0f}",
        'Localización': ci.get('location', 'concentrica'),
        'Lado lindero': ci.get('side', ''),
        'Esquina lindero': ci.get('corner', ''),
        'Mpx': ci.get('mpx') if ci.get('mpx', 0.0) != 0.0 else _ref_mpx,
        'Mpy': ci.get('mpy') if ci.get('mpy', 0.0) != 0.0 else _ref_mpy,
        'Vux': ci.get('vux') if ci.get('vux', 0.0) != 0.0 else _ref_vux,
        'Vuy': ci.get('vuy') if ci.get('vuy', 0.0) != 0.0 else _ref_vuy,
    })

df_edit = pd.DataFrame(edit_rows)

with st.form("class_form", clear_on_submit=False):
    edited = st.data_editor(
        df_edit,
        column_config={
            'Nodo': st.column_config.TextColumn(disabled=True, width='small'),
            'Familia': st.column_config.TextColumn(disabled=True, width='small'),
            'Geometría': st.column_config.TextColumn(disabled=True, width='small'),
            'Localización': st.column_config.SelectboxColumn(options=loc_options, required=True),
            'Lado lindero': st.column_config.SelectboxColumn(options=side_options),
            'Esquina lindero': st.column_config.SelectboxColumn(options=corner_options),
            'Mpx': st.column_config.NumberColumn("Mp en X", min_value=0.0, step=0.01, format="%.2f", width='small'),
            'Mpy': st.column_config.NumberColumn("Mp en Y", min_value=0.0, step=0.01, format="%.2f", width='small'),
            'Vux': st.column_config.NumberColumn("Vu en X", min_value=0.0, step=0.01, format="%.2f", width='small'),
            'Vuy': st.column_config.NumberColumn("Vu en Y", min_value=0.0, step=0.01, format="%.2f", width='small'),
        },
        use_container_width=True,
        hide_index=True,
        key="class_editor",
    )

    apply_class_btn = st.form_submit_button(
        "✅ Aplicar clasificación",
        type="primary",
        use_container_width=True
    )

if apply_class_btn:
    new_cl = {}
    warnings_cl = []
    for i, c in enumerate(design_entities):
        jid = str(c['joint'])
        row = edited.iloc[i]
        loc_val = row['Localización']
        side_val = row['Lado lindero'] if loc_val == 'medianera' else ''
        corner_val = row['Esquina lindero'] if loc_val == 'esquinera' else ''
        mpx = float(row.get('Mpx', 0.0) or 0.0)
        mpy = float(row.get('Mpy', 0.0) or 0.0)
        vux = float(row.get('Vux', 0.0) or 0.0)
        vuy = float(row.get('Vuy', 0.0) or 0.0)

        new_cl[jid] = {
            'location': loc_val,
            'side': side_val,
            'corner': corner_val,
            'ecc_x': 0.0,
            'ecc_y': 0.0,
            'ecc_dir': 'ambas',
            'mpx': mpx,
            'mpy': mpy,
            'vux': vux,
            'vuy': vuy,
        }

        if loc_val == 'medianera' and not side_val:
            warnings_cl.append(f"J{jid}: Medianera sin lado definido")
        if loc_val == 'esquinera' and not corner_val:
            warnings_cl.append(f"J{jid}: Esquinera sin esquina definida")

    if warnings_cl:
        for w in warnings_cl:
            st.error(f"⚠️ {w}")
        st.error("Corrija los campos antes de aplicar.")
    else:
        st.session_state['classifications'] = new_cl
        st.session_state['class_applied'] = True
        ties_auto = deduce_tie_beams(design_entities, new_cl)
        tie_table = ties_dict_to_table(ties_auto, design_entities)
        st.session_state['tie_beams_table'] = tie_table
        st.session_state['tie_version'] = st.session_state.get('tie_version', 0) + 1
        st.session_state.pop('ties_applied', None)
        st.session_state.pop('results', None)
        st.rerun()

# ════════════════════════════════════════════════
# PASO 2: VIGAS DE ENLACE (solo después de aplicar clasificación)
# ════════════════════════════════════════════════
if not st.session_state.get('class_applied'):
    st.info("Configure la clasificación y presione 'Aplicar clasificación' para continuar.")
    st.stop()

cl = st.session_state['classifications']
st.header("🔗 Paso 2 — Vigas de enlace")

# ── Gráfico preliminar: columnas + zapatas conceptuales + vigas ──
def draw_preliminary_plot(columns, cl, tie_rows):
    """Genera el gráfico de planta preliminar."""
    fig = go.Figure()
    col_map = {c['joint']: c for c in columns}

    # Dibujar zapatas conceptuales y columnas
    for c in design_entities:
        jid = c['joint']
        ci = cl.get(jid, {})
        loc = ci.get('location', 'concentrica')
        bx, by = c['bx'], c['by']
        zx, zy, dim = get_preliminary_position(c, ci)
        border, fill = LOC_COLORS.get(loc, LOC_COLORS['concentrica'])

        # Zapata conceptual
        fig.add_shape(type="rect",
            x0=zx-dim/2, y0=zy-dim/2, x1=zx+dim/2, y1=zy+dim/2,
            line=dict(color=border, width=1.5, dash='dash'), fillcolor=fill)
        # Columna
        fig.add_shape(type="rect",
            x0=c['x']-bx/2, y0=c['y']-by/2, x1=c['x']+bx/2, y1=c['y']+by/2,
            line=dict(color='#374151', width=1.5), fillcolor='#9ca3af')
        # Etiqueta
        label = f"<b>J{jid}</b><br>{loc[:5]}"
        fig.add_annotation(x=c['x'], y=c['y']+max(dim/2, by/2)+0.15,
            text=label, showarrow=False, font=dict(size=8, color='#374151'))

    # Dibujar vigas de enlace
    drawn = set()

    def normalize_joint_label(val):
        s = str(val).strip()
        if s == 'Ninguno':
            return s
        if s.startswith('J') and len(s) > 1 and s[1:].isdigit():
            return s[1:]
        return s
    
    for row in tie_rows:
        nodo_id = normalize_joint_label(row['Nodo'])
        col_from = col_map.get(nodo_id)
        if not col_from:
            continue

        for dir_key, connect_key in [('X', 'Conecta_X'), ('Y', 'Conecta_Y')]:
            target = normalize_joint_label(row.get(connect_key, 'Ninguno'))
            if target == 'Ninguno' or target not in col_map:
                continue

            col_to = col_map[target]
            pair = tuple(sorted([nodo_id, target])) + (dir_key,)
            if pair in drawn:
                continue
            drawn.add(pair)

            is_manual = row.get('Origen', 'auto') == 'manual'
            color = '#10b981' if is_manual else '#f97316'
            dash = 'solid' if is_manual else 'dot'

            fig.add_trace(go.Scatter(
                x=[col_from['x'], col_to['x']],
                y=[col_from['y'], col_to['y']],
                mode='lines',
                line=dict(color=color, width=2.5, dash=dash),
                showlegend=False,
                hovertext=f"Viga {dir_key}: {row['Nodo']} → {row.get(connect_key, 'Ninguno')}",
                hoverinfo='text',
            ))

    fig.update_layout(title="Vista preliminar — Zapatas conceptuales + Vigas de enlace")
    apply_theme(fig, height=500)
    return fig

tie_table = st.session_state.get('tie_beams_table', [])
fig_pre = draw_preliminary_plot(design_entities, cl, tie_table)
st.plotly_chart(fig_pre, use_container_width=True)

# ── Sprint A2 Fase 6: Nota sobre zapatas combinadas ──────────────────────────
# Si ya existe un resultado de diseño con zapatas combinadas, mostrar a qué
# nudo combinado pertenece cada columna, para que el usuario sepa que al
# seleccionar J5 como destino de viga, el enlace irá a ZC-01 (que contiene J5+J8).
_prev_final = (st.session_state.get('results') or {}).get('final_footings', [])
_combined_map = {}   # joint_str → footing_id  (solo zapatas combinadas)
for _f in _prev_final:
    if _f.get('type') == 'combined':
        for _jj in _f.get('joint', '').split('+'):
            _jj = _jj.strip()
            if _jj:
                _combined_map[_jj] = _f['id']

if _combined_map:
    _comb_lines = []
    for _fid in sorted(set(_combined_map.values())):
        _joints_in = sorted(j for j, fid in _combined_map.items() if fid == _fid)
        _comb_lines.append(f"**{_fid}**: " + " + ".join(f"J{j}" for j in _joints_in))
    st.info(
        "ℹ️ **Zapatas combinadas del último diseño** — al seleccionar un nudo destino que "
        "pertenece a una combinada, la viga de enlace se conectará a esa zapata completa:\n\n" +
        "  \n".join(_comb_lines)
    )

# ── Tabla editable de vigas ──
st.subheader("Conexiones de vigas de enlace")
if tie_table:
    st.caption("Modifique las conexiones o marque **🗑️ Eliminar** para quitar una fila. Puede agregar vigas abajo.")
else:
    st.caption("No se proponen vigas automáticas. Puede agregar vigas manuales abajo.")

if tie_table:
    # Añadir columna Eliminar si no existe
    _tie_rows_display = []
    for _r in tie_table:
        _tr = dict(_r)
        _tr.setdefault('Eliminar', False)
        _tie_rows_display.append(_tr)
    df_ties = pd.DataFrame(_tie_rows_display)
    tie_key = f"tie_editor_{st.session_state.get('tie_version', 0)}"
    edited_ties = st.data_editor(
        df_ties,
        column_config={
            'Nodo': st.column_config.TextColumn(disabled=True, width='small'),
            'Conecta_X': st.column_config.SelectboxColumn(
                "Conecta en X", options=node_options, required=True, width='small'),
            'Conecta_Y': st.column_config.SelectboxColumn(
                "Conecta en Y", options=node_options, required=True, width='small'),
            'Origen': st.column_config.TextColumn(disabled=True, width='small'),
            'Eliminar': st.column_config.CheckboxColumn("🗑️", width='small'),
        },
        use_container_width=True, hide_index=True, key=tie_key,
    )
else:
    edited_ties = pd.DataFrame(columns=['Nodo', 'Conecta_X', 'Conecta_Y', 'Origen', 'Eliminar'])

# ── Agregar viga manual ──
st.subheader("➕ Agregar viga manual")
ac1, ac2, ac3, ac4 = st.columns([2, 2, 1, 1])
with ac1:
    add_from = st.selectbox("Desde nodo", node_options[1:], key="add_from")
with ac2:
    add_to = st.selectbox("Hacia nodo", node_options[1:], key="add_to")
with ac3:
    add_dir = st.selectbox("Dirección", ['X', 'Y'], key="add_dir")
with ac4:
    st.write("")  # spacer
    if st.button("➕ Agregar", key="btn_add_tie"):
        if add_from == add_to:
            st.error("No puede conectar un nodo consigo mismo.")
        else:
            current_table = st.session_state.get('tie_beams_table', [])
            # Check si el nodo ya tiene fila
            nodo_id = add_from
            existing = [r for r in current_table if r['Nodo'] == nodo_id]
            if existing:
                # Actualizar la conexión en la dirección correspondiente
                for r in current_table:
                    if r['Nodo'] == nodo_id:
                        if add_dir == 'X':
                            r['Conecta_X'] = add_to
                        else:
                            r['Conecta_Y'] = add_to
                        if r['Origen'] == 'auto':
                            r['Origen'] = 'auto+manual'
                        break
            else:
                new_row = {
                    'Nodo': nodo_id,
                    'Conecta_X': add_to if add_dir == 'X' else 'Ninguno',
                    'Conecta_Y': add_to if add_dir == 'Y' else 'Ninguno',
                    'Origen': 'manual',
                }
                current_table.append(new_row)
            st.session_state['tie_beams_table'] = current_table
            st.session_state['tie_version'] = st.session_state.get('tie_version', 0) + 1
            st.session_state.pop('ties_applied', None)
            st.session_state.pop('results', None)
            st.rerun()

# ── Botón APLICAR vigas ──
st.divider()
if st.button("✅ Aplicar vigas y actualizar gráfico", type="primary", use_container_width=True):
    # Leer la tabla editada; descartar filas marcadas con Eliminar=True
    final_tie_rows = []
    if not edited_ties.empty:
        for _, row in edited_ties.iterrows():
            if row.get('Eliminar', False):
                continue   # fila marcada para borrar → se omite
            final_tie_rows.append({
                'Nodo': row['Nodo'],
                'Conecta_X': row['Conecta_X'],
                'Conecta_Y': row['Conecta_Y'],
                'Origen': row.get('Origen', 'auto'),
            })
    st.session_state['tie_beams_table'] = final_tie_rows
    st.session_state['tie_version'] = st.session_state.get('tie_version', 0) + 1
    st.session_state['ties_applied'] = True
    st.session_state.pop('results', None)
    st.rerun()

# ════════════════════════════════════════════════
# PASO 3-5: EJECUTAR DISEÑO
# ════════════════════════════════════════════════
st.divider()
st.header("🚀 Paso 3 — Ejecutar diseño")

if st.session_state.get('ties_applied'):
    st.success("✅ Vigas configuradas. Listo para ejecutar diseño.")
else:
    st.info("Revise las vigas de enlace y presione 'Aplicar vigas' antes de ejecutar el diseño.")

if st.button("🚀 Ejecutar diseño completo", type="primary", use_container_width=True):
    tie_table_final = st.session_state.get('tie_beams_table', [])
    user_ties = table_to_ties_dict(tie_table_final, design_entities)

    with st.spinner("Optimizando geometría y verificando..."):
        md_run = dict(md)

        if md.get("basis_mode") == "support_reactions":
            df_react = st.session_state.get("reactions_df")
            if df_react is None:
                st.error("No se encontró el Excel de reacciones cargado.")
                st.stop()

            md_run["reactions_df"] = df_react

        results = run_design(md_run, cl, params, user_ties=user_ties)

        st.session_state['results'] = results

        # Guardar dimensiones del diseño como base para verificación
        dims_base = {}
        for f in results['final_footings']:
            dims_base[f['id']] = {'B': f['B'], 'L': f['L'], 'h': f['h']}
        st.session_state['dims_design'] = dims_base
        st.session_state.pop('results_verify', None)

        st.rerun()
if 'results' not in st.session_state:
    st.stop()

# ════════════════════════════════════════════════
# MODO VERIFICACIÓN — Editar dimensiones
# ════════════════════════════════════════════════
results = st.session_state['results']
final = results['final_footings']

# Si hay resultados de verificación, usar esos
if 'results_verify' in st.session_state:
    results = st.session_state['results_verify']
    final = results['final_footings']

st.header("📊 Resultados" + (" — Modo verificación" if results.get('verify_mode') else ""))

mc1, mc2, mc3, mc4, mc5 = st.columns(5)
mc1.metric("Cimientos", len(final))
mc2.metric("Combos", f"{results['n_combos_ads']}+{results['n_combos_lrfd']}")
mc3.metric("Solapes", len(results['overlaps']))
mc4.metric("Área", f"{results['total_area']} m²")
mc5.metric("Volumen", f"{results['total_volume']} m³")

# Convergence info
conv = results.get('convergence', {})
if conv:
    conv_txt = f"Convergencia: {conv.get('iterations', 1)} iter"
    if conv.get('resized_footings'):
        conv_txt += f" | Redimensionadas: {', '.join(conv['resized_footings'])}"
    if conv.get('converged'):
        st.success(f"✅ {conv_txt}")
    else:
        st.warning(f"⚠️ {conv_txt} — No convergió completamente")

# ── Tabla editable de verificación ──
with st.expander("✏️ Modificar dimensiones y re-verificar", expanded=False):
    st.caption(
        "Modifique B, L, h y presione **Re-verificar**. "
        "Las zapatas con **🔒** en B o L tienen esa dimensión bloqueada por linderos y no pueden modificarse. "
        "Las marcadas con **📌** tienen una cara libre: la modificación se aplica desde el lindero fijo."
    )
    verify_rows = []
    for f in final:
        is_comb = f.get('type') == 'combined'
        cc = f.get('combined_constraints', {}) if is_comb else {}
        locked_B = cc.get('locked_B', False)
        locked_L = cc.get('locked_L', False)
        pinned_x = cc.get('pinned_x', False)
        pinned_y = cc.get('pinned_y', False)
        # Etiqueta de restricción
        rest_b = '🔒 bloqueada' if locked_B else ('📌 una cara' if pinned_x else '↔ libre')
        rest_l = '🔒 bloqueada' if locked_L else ('📌 una cara' if pinned_y else '↔ libre')
        verify_rows.append({
            'ID': f['id'],
            'Tipo': f.get('type', '?')[:5],
            'B': f['B'],
            'Rest. B': rest_b,
            'L': f['L'],
            'Rest. L': rest_l,
            'h': f['h'],
            'Estado': f['st'],
        })
    df_verify = pd.DataFrame(verify_rows)
    ver_key = f"verify_editor_{st.session_state.get('_verify_v', 0)}"
    edited_dims = st.data_editor(
        df_verify,
        column_config={
            'ID':     st.column_config.TextColumn(disabled=True, width='small'),
            'Tipo':   st.column_config.TextColumn(disabled=True, width='small'),
            'B':      st.column_config.NumberColumn(min_value=0.30, step=0.05, format="%.2f"),
            'Rest. B':st.column_config.TextColumn(disabled=True, width='small'),
            'L':      st.column_config.NumberColumn(min_value=0.30, step=0.05, format="%.2f"),
            'Rest. L':st.column_config.TextColumn(disabled=True, width='small'),
            'h':      st.column_config.NumberColumn(min_value=0.20, step=0.05, format="%.2f"),
            'Estado': st.column_config.TextColumn(disabled=True, width='small'),
        },
        use_container_width=True, hide_index=True, key=ver_key,
    )
    if st.button("🔄 Re-verificar con dimensiones modificadas", type="secondary", use_container_width=True):
        user_dims = {}
        # Build lookup: footing id → combined_constraints (for combined footings)
        _cc_map = {f['id']: f.get('combined_constraints', {}) for f in final if f.get('type') == 'combined'}
        for _, row in edited_dims.iterrows():
            fid = str(row['ID'])
            cc_row = _cc_map.get(fid, {})
            B_usr = float(row['B']); L_usr = float(row['L']); h_usr = float(row['h'])
            if row['Tipo'][:5] == 'combi':
                # Para combined: respetar bloqueos de lindero
                user_dims[fid] = {
                    'B': B_usr if not cc_row.get('locked_B') else None,
                    'L': L_usr if not cc_row.get('locked_L') else None,
                    'h': h_usr,
                }
            else:
                user_dims[fid] = {'B': B_usr, 'L': L_usr, 'h': h_usr}
        tie_table_final = st.session_state.get('tie_beams_table', [])
        user_ties = table_to_ties_dict(tie_table_final, design_entities)
        with st.spinner("Re-verificando con dimensiones del usuario..."):
            results_v = run_design(md, cl, params, user_ties=user_ties, user_dims=user_dims)
            st.session_state['results_verify'] = results_v
            st.session_state['_verify_v'] = st.session_state.get('_verify_v', 0) + 1
            st.rerun()

# Status styling
def sty(val):
    if val in ('PRELIMINAR_OK','PRELIMINAR_OK_COMBINADA'):
        return 'background-color:#d1fae5;color:#065f46'
    if 'REVISION' in str(val):
        return 'background-color:#dbeafe;color:#1e40af'
    if val == 'REVISAR_h':
        return 'background-color:#fef3c7;color:#92400e'
    if val == 'NO_CUMPLE':
        return 'background-color:#fee2e2;color:#991b1b'
    return ''

tab_plan, tab_resumen, tab_aisladas, tab_combinadas, tab_vigas_res, tab_audit, tab_audit_v, tab_export = st.tabs([
    "📐 Planta", "📋 Resumen ejecutivo", "📊 Zapatas aisladas", "📊 Zapatas combinadas",
    "📊 Sistemas de vigas", "🔍 Aud. Zapatas", "🔗 Aud. Vigas", "📥 Exportar"
])

# ════════════════════════════════════════════════
# TAB: PLANTA DE RESULTADOS
# ════════════════════════════════════════════════
with tab_plan:
    fig = go.Figure()
    ties_data = st.session_state.get('tie_beams_table', [])
    col_map = {c['joint']: c for c in columns}

    for f in final:
        fx = f.get('x_footing', f['x']); fy_p = f.get('y_footing', f['y'])
        B, L = f['B'], f['L']
        border, fill = STATUS_COLORS.get(f['st'], ('#059669', 'rgba(5,150,105,0.10)'))
        if f.get('type') == 'combined':
            fill = 'rgba(219,39,119,0.08)'

        # Zapata
        fig.add_shape(type="rect", x0=fx-B/2, y0=fy_p-L/2, x1=fx+B/2, y1=fy_p+L/2,
            line=dict(color=border, width=2), fillcolor=fill)

        # Columnas dentro de la zapata
        cont_details = f.get('containment', {}).get('details', [])
        cont_map = {d['joint']: d['contained'] for d in cont_details} if cont_details else {}
        for col in f.get('cols', []):
            col_x = col.get('x', fx); col_y = col.get('y', fy_p)
            bx, by = col.get('bx', 0.15), col.get('by', 0.15)
            col_ok = cont_map.get(col['joint'], True)
            col_color = '#dc2626' if not col_ok else '#4b5563'
            fig.add_shape(type="rect",
                x0=col_x-bx/2, y0=col_y-by/2, x1=col_x+bx/2, y1=col_y+by/2,
                line=dict(color=col_color, width=1.5 if col_ok else 2), fillcolor='#9ca3af')

        # Etiqueta
        loc_str = f.get('classification', {}).get('location', '')[:6] if f.get('type') != 'combined' else f.get('scheme', '')[:8]
        fig.add_annotation(x=fx, y=fy_p+L/2+0.15,
            text=f"<b>{f['id']}</b><br>{B}×{L} h={f['h']}<br>{loc_str}",
            showarrow=False, font=dict(size=8, color='#374151'))

    # Vigas de enlace
    drawn_ties = set()
    for row in ties_data:
        nodo_id = str(row['Nodo']).replace('J', '')
        col_from = col_map.get(nodo_id)
        if not col_from:
            continue
        for connect_key in ['Conecta_X', 'Conecta_Y']:
            target = str(row.get(connect_key, 'Ninguno')).replace('J', '')
            if target == 'Ninguno' or target not in col_map:
                continue
            col_to = col_map[target]
            pair = tuple(sorted([nodo_id, target]))
            if pair in drawn_ties:
                continue
            drawn_ties.add(pair)
            fig.add_trace(go.Scatter(
                x=[col_from['x'], col_to['x']], y=[col_from['y'], col_to['y']],
                mode='lines', line=dict(color='#f97316', width=2, dash='dot'),
                showlegend=False))

    fig.update_layout(title="Planta de cimentaciones — Resultados")
    apply_theme(fig, height=600)
    st.plotly_chart(fig, use_container_width=True)

# ════════════════════════════════════════════════
# TAB: RESUMEN EJECUTIVO
# ════════════════════════════════════════════════
with tab_resumen:
    st.subheader("Resumen por zapata — Combo controlante ASD + LRFD")
    res_rows = []
    for f in final:
        rx = propose_rebar(f['Asx'], f['B'], rec)
        ry = propose_rebar(f['Asy'], f['L'], rec)
        loc_info = f.get('classification', {}).get('location', '') if f.get('type') != 'combined' else f.get('scheme', '')
        # Combo ASD crítico
        ads_sorted = sorted(f.get('ads_audit', []), key=lambda a: a['ratio'], reverse=True)
        ctrl_ads = ads_sorted[0] if ads_sorted else {}
        # Combo LRFD crítico
        lrfd_sorted = sorted(f.get('lrfd_audit', []), key=lambda a: a.get('punch_ratio', 0), reverse=True)
        ctrl_lrfd = lrfd_sorted[0] if lrfd_sorted else {}
        res_rows.append({
            'ID': f['id'], 'Tipo': f.get('type', '?')[:5], 'Loc': loc_info[:10],
            'B': f['B'], 'L': f['L'], 'h': f['h'],
            'ASD ctrl': ctrl_ads.get('combo', '')[:25],
            'Ratio': ctrl_ads.get('ratio', 0),
            'qmax': f['qmax'],
            'q1': ctrl_ads.get('q1', 0), 'q2': ctrl_ads.get('q2', 0),
            'q3': ctrl_ads.get('q3', 0), 'q4': ctrl_ads.get('q4', 0),
            'FS volc': f.get('fs_volc_min', 999),
            'FS desl': f.get('fs_desl_min', 999),
            'LRFD ctrl': ctrl_lrfd.get('combo', '')[:25],
            'Pu': f['Pu'], 'P%': round(f['pr']*100, 1), 'V%': round(f['sr']*100, 1),
            'Asx': f['Asx'], 'Ref X': rx['text'],
            'Asy': f['Asy'], 'Ref Y': ry['text'],
            'Estado': f['st'],
        })
    df_res = pd.DataFrame(res_rows)
    def sty_res(val):
        if isinstance(val, str):
            return sty(val)
        return ''
    st.dataframe(df_res.style.map(sty_res, subset=['Estado']),
                 use_container_width=True, hide_index=True, height=min(600, 40+35*len(res_rows)))

    # Resumen de vigas
    valid_sys = [s for s in results.get('tie_systems', [])
                 if s.get('status') not in ('insuficiente', 'distancia_insuficiente')]
    if valid_sys:
        st.divider()
        st.subheader("Resumen sistemas de vigas — Combo controlante")
        vr = []
        for sys in valid_sys:
            vr.append({
                'Sistema': sys['system_id'], 'Dir': sys['direction'],
                'Zapatas': ', '.join(sys.get('footings', [])),
                'L(m)': sys['total_length'],
                'b×h(cm)': f"{sys['b_viga']*100:.0f}×{sys['h_viga']*100:.0f}",
                'ADS ctrl': sys.get('ads_control', '')[:25],
                'LRFD ctrl': sys.get('lrfd_control', '')[:25],
                'Mu+': sys['Mu_max_pos'], 'Mu-': sys['Mu_max_neg'],
                'Vu': sys['Vu_max'], 'V%': f"{sys['sr_viga']*100:.1f}",
                'As inf': sys['As_inf_text'], 'As sup': sys['As_sup_text'],
                'Estado': sys['status'],
            })
        st.dataframe(pd.DataFrame(vr), use_container_width=True, hide_index=True)

# ════════════════════════════════════════════════
# TAB: ZAPATAS AISLADAS
# ════════════════════════════════════════════════
with tab_aisladas:
    isolated = [f for f in final if f.get('type') != 'combined']
    if not isolated:
        st.info("No hay zapatas aisladas.")
    else:
        rows = []
        for f in isolated:
            rx = propose_rebar(f['Asx'], f['B'], rec)
            ry = propose_rebar(f['Asy'], f['L'], rec)
            loc_info = f.get('classification', {}).get('location', '')
            # Caso crítico LRFD (mayor ratio de punzonamiento)
            _lc = max(f.get('lrfd_audit', [{}]), key=lambda a: a.get('punch_ratio', 0), default={})
            rows.append({
                'ID': f['id'], 'Loc': loc_info[:8],
                'B': f['B'], 'L': f['L'], 'h': f['h'],
                'qmax': f['qmax'], 'qmin': f['qmin'],
                'Pu': f['Pu'],
                'Mux*': _lc.get('Mux', 0), 'Muy*': _lc.get('Muy', 0),
                'ex*': _lc.get('ex', 0), 'ey*': _lc.get('ey', 0),
                'Vu_x': _lc.get('Vu_x', 0), 'Vx%': round(_lc.get('sr_x', 0)*100, 1),
                'Vu_y': _lc.get('Vu_y', 0), 'Vy%': round(_lc.get('sr_y', 0)*100, 1),
                'Vu_pun': _lc.get('Vu_punch', 0), 'Pun%': round(_lc.get('punch_ratio', 0)*100, 1),
                'Asx': f['Asx'], 'Ref X': rx['text'],
                'Asy': f['Asy'], 'Ref Y': ry['text'],
                'Estado': f['st'],
            })
        df_iso = pd.DataFrame(rows)
        st.dataframe(df_iso.style.map(sty, subset=['Estado']),
                     use_container_width=True, hide_index=True)
        st.caption(f"{len(isolated)} zapatas aisladas")

# ════════════════════════════════════════════════
# TAB: ZAPATAS COMBINADAS
# ════════════════════════════════════════════════
with tab_combinadas:
    combined = [f for f in final if f.get('type') == 'combined']
    if not combined:
        st.info("No hay zapatas combinadas (sin solapamientos detectados).")
    else:
        rows = []
        for f in combined:
            rx = propose_rebar(f['Asx'], f['B'], rec)
            ry = propose_rebar(f['Asy'], f['L'], rec)
            cols_txt = '+'.join(c['joint'] for c in f.get('cols', []))
            _lc = max(f.get('lrfd_audit', [{}]), key=lambda a: a.get('punch_ratio', 0), default={})
            rows.append({
                'ID': f['id'], 'Esquema': f.get('scheme', '')[:12],
                'Columnas': cols_txt,
                'B': f['B'], 'L': f['L'], 'h': f['h'],
                'qmax': f['qmax'], 'qmin': f['qmin'],
                'Pu': f['Pu'],
                'Mux*': _lc.get('Mux', 0), 'Muy*': _lc.get('Muy', 0),
                'ex*': _lc.get('ex', 0), 'ey*': _lc.get('ey', 0),
                'Vu_x': _lc.get('Vu_x', 0), 'Vx%': round(_lc.get('sr_x', 0)*100, 1),
                'Vu_y': _lc.get('Vu_y', 0), 'Vy%': round(_lc.get('sr_y', 0)*100, 1),
                'Vu_pun': _lc.get('Vu_punch', 0), 'Pun%': round(_lc.get('punch_ratio', 0)*100, 1),
                'Asx': f['Asx'], 'Ref X': rx['text'],
                'Asy': f['Asy'], 'Ref Y': ry['text'],
                'As sup': f.get('As_long_top', 0),
                'Estado': f['st'],
            })
        df_comb = pd.DataFrame(rows)
        st.dataframe(df_comb.style.map(sty, subset=['Estado']),
                     use_container_width=True, hide_index=True)
        st.caption(f"{len(combined)} zapatas combinadas")

# ════════════════════════════════════════════════
# TAB: SISTEMAS DE VIGAS
# ════════════════════════════════════════════════
with tab_vigas_res:
    valid_sys = [s for s in results.get('tie_systems', [])
                 if s.get('status') not in ('insuficiente', 'distancia_insuficiente')]
    if not valid_sys:
        st.info("No hay sistemas de vigas de enlace.")
    else:
        rows = []
        for sys in valid_sys:
            _tipo = '🔗 intra' if sys.get('intra_combined') else '↔ enlace'
            rows.append({
                'Sistema': sys['system_id'],
                'Tipo': _tipo,
                'Dir': sys['direction'],
                'Zapatas': ', '.join(sys.get('footings', [])),
                'Nodos': sys['num_nodes'],
                'L total': f"{sys['total_length']}m",
                'Sección': f"{sys['b_viga']*100:.0f}×{sys['h_viga']*100:.0f}cm",
                'Mu+': f"{sys['Mu_max_pos']} kN·m",
                'Mu-': f"{sys['Mu_max_neg']} kN·m",
                'Vu': f"{sys['Vu_max']} kN",
                'As inf': f"{sys['As_inf']}cm² → {sys['As_inf_text']}",
                'As sup': f"{sys['As_sup']}cm² → {sys['As_sup_text']}",
                'V%': f"{sys['sr_viga']*100:.1f}%",
                'Estado': sys['status'],
            })
        df_sys = pd.DataFrame(rows)
        st.dataframe(df_sys, use_container_width=True, hide_index=True)
        st.caption(f"{len(valid_sys)} sistemas de vigas")

# ════════════════════════════════════════════════
# TAB: AUDITORÍA ZAPATAS
# ════════════════════════════════════════════════
with tab_audit:
    sel_id = st.selectbox("Seleccionar cimiento:", [f['id'] for f in final], key="aud_sel")
    sf = next(f for f in final if f['id'] == sel_id)
    rx = propose_rebar(sf['Asx'], sf['B'], rec)
    ry = propose_rebar(sf['Asy'], sf['L'], rec)

    amc1, amc2, amc3, amc4 = st.columns(4)
    amc1.metric("Dimensiones", f"{sf['B']}×{sf['L']}×{sf['h']}m")
    amc2.metric("Pu máx", f"{sf['Pu']} kN")
    amc3.metric("Punzonamiento", f"{sf['pr']*100:.1f}%")
    amc4.metric("Estado", sf['st'])

    # FS volcamiento y deslizamiento
    fs_v = sf.get('fs_volc_min', 999); fs_d = sf.get('fs_desl_min', 999)
    fsc1, fsc2, fsc3, fsc4 = st.columns(4)
    fsc1.metric("FS Volc. mín", f"{fs_v:.1f}" if fs_v < 900 else "∞",
                delta="OK" if not sf.get('any_volc_fail') else "NO CUMPLE",
                delta_color="normal" if not sf.get('any_volc_fail') else "inverse")
    fsc2.metric("FS Desl. mín", f"{fs_d:.1f}" if fs_d < 900 else "∞",
                delta="OK" if not sf.get('any_desl_fail') else "NO CUMPLE",
                delta_color="normal" if not sf.get('any_desl_fail') else "inverse")
    fsc3.metric("qmax", f"{sf['qmax']} kPa")
    fsc4.metric("qmin", f"{sf['qmin']} kPa")

    st.markdown(f"**Refuerzo:** Asx={sf['Asx']}cm² → **{rx['text']}** | Asy={sf['Asy']}cm² → **{ry['text']}**")

    sch = sf.get('scheme', '?')
    egx = sf.get('e_geo_x', 0); egy = sf.get('e_geo_y', 0)
    if sch != 'aislada' or egx > 0.01 or egy > 0.01:
        st.info(f"**Esquema:** {sch} | e_geo_x={egx}m e_geo_y={egy}m")

    # Restricciones heredadas (combinadas)
    cc = sf.get('combined_constraints', {})
    if cc:
        parts = []
        if cc.get('has_edge_constraint'): parts.append(f"Bordes: {cc.get('edge_sides', [])}")
        if cc.get('has_corner_constraint'): parts.append(f"Esquinas: {cc.get('corners', [])}")
        if parts:
            st.warning(f"⚠️ Restricciones heredadas: {' | '.join(parts)}")
        # Mostrar estado de bloqueo dimensional
        _lock_msgs = []
        if cc.get('locked_B'):
            lp = cc.get('left_pin'); rp = cc.get('right_pin')
            _lock_msgs.append(f"🔒 **B bloqueada** por linderos X— y X+ (B fija = {round(rp-lp,2) if lp and rp else '?'}m). No se puede ampliar.")
        if cc.get('locked_L'):
            bp = cc.get('bot_pin'); tp = cc.get('top_pin')
            _lock_msgs.append(f"🔒 **L bloqueada** por linderos Y— y Y+ (L fija = {round(tp-bp,2) if bp and tp else '?'}m). No se puede ampliar.")
        if not cc.get('locked_B') and cc.get('pinned_x'):
            _lock_msgs.append(f"📌 B anclada en una cara — modificación se aplica hacia el lado libre.")
        if not cc.get('locked_L') and cc.get('pinned_y'):
            _lock_msgs.append(f"📌 L anclada en una cara — modificación se aplica hacia el lado libre.")
        for _lm in _lock_msgs:
            st.info(_lm)
        cont = sf.get('containment', {})
        if cont.get('details'):
            not_ok = [d for d in cont['details'] if not d['contained']]
            if not_ok:
                st.error(f"❌ Columnas NO contenidas (linderos impiden expansión): "
                         f"{', '.join('J'+d['joint'] for d in not_ok)}")
            else:
                st.success(f"✅ Todas las columnas contenidas")

    # Jerarquía de dimensiones (zapatas combinadas)
    _hcheck = sf.get('hierarchy_check')
    if sf.get('type') == 'combined' and _hcheck:
        st.divider()
        st.subheader("📐 Jerarquía de dimensiones")
        _hc1, _hc2 = st.columns(2)
        _hc1.metric("B mín. por columnas aisladas", f"{_hcheck.get('B_min_hier', '?')} m",
                    delta=f"{round(sf['B'] - _hcheck.get('B_min_hier', sf['B']), 3):+.3f} m vs actual")
        _hc2.metric("L mín. por columnas aisladas", f"{_hcheck.get('L_min_hier', '?')} m",
                    delta=f"{round(sf['L'] - _hcheck.get('L_min_hier', sf['L']), 3):+.3f} m vs actual")
        _b_isos = _hcheck.get('B_iso_list', [])
        _l_isos = _hcheck.get('L_iso_list', [])
        if _b_isos:
            st.caption(f"B por columna: {' / '.join(f'{v}m' for v in _b_isos)}  →  max = {_hcheck.get('B_min_hier')}m")
        if _l_isos:
            st.caption(f"L por columna: {' / '.join(f'{v}m' for v in _l_isos)}  →  max = {_hcheck.get('L_min_hier')}m")
        if _hcheck.get('ok'):
            st.success("✅ Dimensiones de zapata combinada satisfacen jerarquía de columnas aisladas.")
        else:
            if _hcheck.get('resized'):
                st.warning(f"🔄 Zapata redimensionada por jerarquía: "
                           f"B {_hcheck.get('B_before','?')} → {sf['B']} m, "
                           f"L {_hcheck.get('L_before','?')} → {sf['L']} m")
            else:
                st.error("⚠️ Dimensiones por debajo del mínimo jerárquico (linderos bloquean expansión).")

    # Sistemas enlazados
    sys_list = results.get('tie_systems', [])
    my_systems = [s for s in sys_list if sf.get('id') in s.get('footings', [])]
    if my_systems:
        st.divider()
        st.subheader("🔗 Sistemas de enlace")
        dp = sf.get('system_dP', 0)
        if dp > 0:
            sc1, sc2, sc3 = st.columns(3)
            sc1.metric("ΔP sistema", f"{dp} kN")
            sc2.metric("qmax c/sistema", f"{sf.get('qmax_with_system', 0)} kPa")
            sc3.metric("Ratio c/sistema", f"{sf.get('ratio_with_system', 0):.3f}")
            if sf.get('needs_resize'):
                st.error("⚠️ Requiere redimensionar con cargas del sistema")

    # Sub-tabs
    if sf.get('type') == 'combined' and sf.get('combined_analysis'):
        t1, t2, t3, t4 = st.tabs(["📦 Cargas", "📐 ADS", "📈 V-M y Acero", "🔩 LRFD"])
    else:
        t1, t2, t3 = st.tabs(["📦 Cargas", "📐 ADS", "🔩 LRFD"])

    with t1:
        cf = sf.get('column_forces', {})
        if cf:
            for jid, pats in cf.items():
                st.subheader(f"Columna J{jid}")
                rows_cf = [{'Caso': pat, 'P': round(v.get('F3', 0), 2), 'Mx': round(v.get('M1', 0), 2),
                    'My': round(v.get('M2', 0), 2), 'Vx': round(v.get('F1', 0), 2), 'Vy': round(v.get('F2', 0), 2)}
                    for pat, v in sorted(pats.items())]
                st.dataframe(pd.DataFrame(rows_cf), use_container_width=True, hide_index=True)

    with t2:
        ads_sorted = sorted(sf.get('ads_audit', []), key=lambda a: a['ratio'], reverse=True)
        show_all = st.checkbox("Mostrar todas", value=False, key="ads_all_a")
        display = ads_sorted if show_all else ads_sorted[:12]
        rows_a = [{'Combo': a['combo'], 'Grupo': a['group'], 'qadm': a['qadm'],
            'P+Wp': a['P_total'], 'Mx': a['Mx'], 'My': a['My'], 'Vx': a.get('Vx', 0), 'Vy': a.get('Vy', 0),
            'ex': a.get('ex', 0), 'ey': a.get('ey', 0), 'Contacto': a.get('contact', '?'),
            'q1': a.get('q1', 0), 'q2': a.get('q2', 0), 'q3': a.get('q3', 0), 'q4': a.get('q4', 0),
            'qmax': a['qmax'], 'qmin': a['qmin'], 'Ratio': a['ratio'],
            'FSv': a.get('fs_volc_min', 999), 'FSd': a.get('fs_desl_min', 999),
            } for a in display]
        dfa = pd.DataFrame(rows_a)
        def cr(v):
            if isinstance(v, (int, float)):
                if v > 1.05: return 'background-color:#fee2e2;color:#991b1b'
                if v > 0.85: return 'background-color:#fef3c7;color:#92400e'
            return ''
        def cr_fs(v):
            if isinstance(v, (int, float)) and v < 900:
                if v < 1.5: return 'background-color:#fee2e2;color:#991b1b'
                if v < 2.0: return 'background-color:#fef3c7;color:#92400e'
            return ''
        styled = dfa.style.map(cr, subset=['Ratio']).map(cr_fs, subset=['FSv', 'FSd'])
        st.dataframe(styled, use_container_width=True, hide_index=True)

        # ── Cargas por columna (solo zapatas combinadas) ──────────────────────
        _ca_loads = sf.get('col_ads_loads', [])
        if _ca_loads:
            st.divider()
            st.markdown("**📋 Cargas por columna individual (ADS)**")
            st.caption("Fuerzas en la base de cada columna antes de transportar al centroide de la zapata combinada.")
            _df_cal = pd.DataFrame(_ca_loads)
            # Ordenar: columna → combo → P desc
            if not _df_cal.empty:
                _df_cal = _df_cal.sort_values(['Columna', 'P (kN)'], ascending=[True, False])
            st.dataframe(_df_cal, use_container_width=True, hide_index=True)

    # V-M tab (solo combinadas)
    if sf.get('type') == 'combined' and sf.get('combined_analysis'):
        with t3:
            ca = sf['combined_analysis']; sd = sf.get('steel_diagram')
            st.caption(f"Eje: **{sf.get('longitudinal_axis', '?').upper()}** | w={ca.get('w', 0)} kN/m")
            fig_vm = go.Figure()
            fig_vm.add_trace(go.Scatter(x=ca['stations'], y=ca['V'], name='V(kN)',
                line=dict(color='#dc2626', width=2)))
            fig_vm.add_trace(go.Scatter(x=ca['stations'], y=ca['M'], name='M(kN·m)',
                line=dict(color='#2563eb', width=2)))
            fig_vm.add_hline(y=0, line_dash="dot", line_color="#9ca3af")
            fig_vm.update_layout(title="Diagrama V-M longitudinal")
            apply_theme(fig_vm, height=350, equal_axes=False)
            st.plotly_chart(fig_vm, use_container_width=True, key=f"vm_comb_{sf['id']}")

            if sd and sd.get('stations'):
                fig_as = go.Figure()
                fig_as.add_trace(go.Scatter(x=sd['stations'], y=sd['As_inf'], name='As inf',
                    fill='tozeroy', line=dict(color='#059669', width=2), fillcolor='rgba(5,150,105,0.15)'))
                fig_as.add_trace(go.Scatter(x=sd['stations'], y=[-v for v in sd['As_sup']], name='As sup',
                    fill='tozeroy', line=dict(color='#d97706', width=2), fillcolor='rgba(217,119,6,0.15)'))
                fig_as.add_hline(y=0, line_dash="dot", line_color="#9ca3af")
                fig_as.update_layout(title="Acero requerido [+inf, -sup]")
                apply_theme(fig_as, height=280, equal_axes=False)
                st.plotly_chart(fig_as, use_container_width=True, key=f"as_comb_{sf['id']}")

    # LRFD tab
    lrfd_tab = t4 if (sf.get('type') == 'combined' and sf.get('combined_analysis')) else t3
    with lrfd_tab:
        lrfd_sorted = sorted(sf.get('lrfd_audit', []), key=lambda a: a.get('punch_ratio', 0), reverse=True)
        show_l = st.checkbox("Mostrar todas LRFD", value=False, key="lrfd_all_a")
        display_l = lrfd_sorted if show_l else lrfd_sorted[:10]
        rows_l = [{
            'Combo': a['combo'],
            'Pu (kN)': a['Pu'],
            'Mux (kN·m)': a['Mux'],
            'Muy (kN·m)': a['Muy'],
            'ex (m)': a.get('ex', 0),
            'ey (m)': a.get('ey', 0),
            # ─── Punzonamiento (ACI 318 §22.6) ───
            'Vu_pun (kN)': a.get('Vu_punch', 0),
            'vu_pun (kPa)': a.get('vu_max', 0),
            'φvc_pun (kPa)': a.get('phi_vc', 0),
            'Pun%': round(a.get('punch_ratio', 0) * 100, 1),
            # ─── Cortante 1-vía dirección X ───
            'Vu_x (kN)': a.get('Vu_x', 0),
            'φVn_x (kN)': a.get('phi_Vn_x', 0),
            'Vx%': round(a.get('sr_x', 0) * 100, 1),
            # ─── Cortante 1-vía dirección Y ───
            'Vu_y (kN)': a.get('Vu_y', 0),
            'φVn_y (kN)': a.get('phi_Vn_y', 0),
            'Vy%': round(a.get('sr_y', 0) * 100, 1),
            'Vcorte%': round(a.get('shear_ratio', 0) * 100, 1),
            'Mu_x (kN·m)': a.get('Mu_x', 0),
            'Mu_y (kN·m)': a.get('Mu_y', 0),
            'As_x (cm²)': a.get('As_x', 0),
            'As_y (cm²)': a.get('As_y', 0),
        } for a in display_l]
        st.caption("Ordenado por ratio de punzonamiento (mayor primero). "
                   "Pun% = vu_pun/φvc_pun × 100. Vx%/Vy% = cortante 1-vía por dirección.")
        st.dataframe(pd.DataFrame(rows_l), use_container_width=True, hide_index=True)

        # ── Punzonamiento por columna individual (solo zapatas combinadas) ──────
        _cpa = sf.get('col_punch_audit', [])
        if _cpa:
            st.divider()
            st.markdown("**🔩 Punzonamiento por columna individual (ACI 318 §22.6)**")
            st.caption(
                "Cada columna de la zapata combinada se verifica de forma independiente. "
                "Vu_pun = Pu·(1 − Ap/A)  donde  Ap = (bx+d)·(by+d),  A = B·L (zapata completa). "
                "El cortante 1-vía (viga amplia) se verifica globalmente en la tabla anterior."
            )
            _df_cp = pd.DataFrame(_cpa)
            # Colorear fondo de Pun%: rojo >100, amarillo >75
            if not _df_cp.empty and 'Pun%' in _df_cp.columns:
                def _cr_pun(v):
                    if isinstance(v, (int, float)):
                        if v > 100: return 'background-color:#fee2e2;color:#991b1b'
                        if v > 75:  return 'background-color:#fef3c7;color:#92400e'
                    return ''
                _df_cp = _df_cp.sort_values(['Columna', 'Pun%'], ascending=[True, False])
                st.dataframe(_df_cp.style.map(_cr_pun, subset=['Pun%']),
                             use_container_width=True, hide_index=True)
            else:
                st.dataframe(_df_cp, use_container_width=True, hide_index=True)

# ════════════════════════════════════════════════
# TAB: AUDITORÍA VIGAS
# ════════════════════════════════════════════════
with tab_audit_v:
    valid_sys = [s for s in results.get('tie_systems', [])
                 if s.get('status') not in ('insuficiente', 'distancia_insuficiente')]
    if not valid_sys:
        st.info("No hay sistemas de vigas de enlace.")
    else:
        sys_names = [f"{s['system_id']} ({s['direction']}, {s['num_nodes']}n)" for s in valid_sys]
        sel_sys_idx = st.selectbox("Seleccionar sistema", range(len(sys_names)),
                                   format_func=lambda i: sys_names[i])
        sd = valid_sys[sel_sys_idx]

        rc1, rc2, rc3, rc4 = st.columns(4)
        rc1.metric("Sección", f"{sd['b_viga']*100:.0f}×{sd['h_viga']*100:.0f}cm")
        rc2.metric("Mu+ máx", f"{sd['Mu_max_pos']} kN·m")
        rc3.metric("Vu máx", f"{sd['Vu_max']} kN")
        rc4.metric("Cortante", f"{sd['sr_viga']*100:.0f}%",
                   delta="OK" if sd['sr_viga'] <= 1 else "REVISAR")

        vt1, vt2, vt3, vt4, vt5, vt6 = st.tabs([
            "🏗️ Sistema", "📦 Cargas", "📐 ADS", "🔩 LRFD", "📊 Resultados", "📈 Diagramas"])

        # ── Sistema ──
        with vt1:
            geom = sd.get('geometry', {})
            st.markdown(f"**ID:** {sd['system_id']} | **Dir:** {sd['direction']} | "
                       f"**Nodos:** {sd['num_nodes']} | **L total:** {sd['total_length']}m")
            st.markdown(f"**Apoyos:** {geom.get('supports', [])}m | "
                       f"**Span:** {geom.get('span', 0)}m | **Voladizos:** {geom.get('overhangs', [])}m")

            # Esquema gráfico
            fig_s = go.Figure()
            loads = geom.get('loads', [])
            sups = geom.get('supports', [0, 1])
            x_all = [ld['x_load'] for ld in loads] + [ld['x_sup'] for ld in loads] + sups
            xmin_g = min(x_all) - 0.5; xmax_g = max(x_all) + 0.5

            fig_s.add_trace(go.Scatter(x=[xmin_g, xmax_g], y=[0, 0], mode='lines',
                line=dict(color='#f97316', width=6), showlegend=False))
            for sx in sups:
                fig_s.add_trace(go.Scatter(x=[sx-0.15, sx+0.15, sx, sx-0.15],
                    y=[-0.15, -0.15, 0, -0.15], fill='toself',
                    fillcolor='rgba(37,99,235,0.4)', line=dict(color='#2563eb', width=2), showlegend=False))
                fig_s.add_annotation(x=sx, y=-0.3, text=f"▲ x={sx}m",
                    showarrow=False, font=dict(size=8, color='#2563eb'))
            for ld in loads:
                fig_s.add_annotation(x=ld['x_load'], y=0.9, text=f"↓ P<br><b>{ld['fid']}</b>",
                    showarrow=True, arrowhead=2, arrowcolor='#dc2626', ay=-50,
                    font=dict(size=10, color='#dc2626'))
                if abs(ld['ecc']) > 0.01:
                    fig_s.add_trace(go.Scatter(x=[ld['x_load'], ld['x_sup']], y=[0.05, 0.05],
                        mode='lines+text', line=dict(color='#d97706', width=1, dash='dash'),
                        text=[f"e={ld['ecc']}m", ''], textposition='top center', showlegend=False))
                fig_s.add_shape(type="rect", x0=ld['x_sup']-0.3, y0=-0.5,
                    x1=ld['x_sup']+0.3, y1=-0.15,
                    line=dict(color='#059669', width=1), fillcolor='rgba(5,150,105,0.12)')

            fig_s.update_layout(title=f"Modelo de viga — Dir {sd['direction']}",
                yaxis=dict(showticklabels=False, range=[-0.7, 1.2]),
                xaxis=dict(range=[xmin_g, xmax_g]))
            apply_theme(fig_s, height=280, equal_axes=False)
            st.plotly_chart(fig_s, use_container_width=True, key=f"beam_model_{sd['system_id']}")

            aud = sd.get('audit', {})
            cols_a = st.columns(4)
            cols_a[0].markdown("✅ ΣFy" if aud.get('force_balance_ok') else
                f"❌ ΣFy err={aud.get('force_balance_error', 0)}kN")
            cols_a[1].markdown("✅ ΣM" if aud.get('moment_balance_ok') else
                f"❌ ΣM err={aud.get('moment_balance_error', 0)}kNm")
            cols_a[2].markdown("✅ V coherente" if aud.get('shear_consistent') else
                f"⚠️ V: {sd['Vu_max']}kN vs R={aud.get('max_reaction_lrfd', 0)}kN")
            cols_a[3].markdown("✅ Zapatas OK" if aud.get('footings_updated') else "⚠️ Redimensionar")

        # ── Cargas ──
        with vt2:
            ctrl_combo = sd.get('ads_control', '')
            ctrl_state = next((cs for cs in sd.get('combo_states_ads', [])
                              if cs['combo'] == ctrl_combo), None)
            if ctrl_state:
                st.markdown(f"**Combo ADS controlante:** {ctrl_combo}")
                nd_rows = []
                for i, nd in enumerate(ctrl_state.get('nodes', [])):
                    dr = ctrl_state['delta_R'][i] if i < len(ctrl_state.get('delta_R', [])) else 0
                    nd_rows.append({
                        'Zapata': sd['footings'][i] if i < len(sd['footings']) else '?',
                        'x_col': nd.get('x_load', '?'), 'x_zap': nd.get('x_sup', '?'),
                        'P(kN)': nd['P'], 'Mcol(kNm)': nd['Mcol'], 'P·e(kNm)': nd['Pe'],
                        'ecc(m)': nd['ecc'], 'Msys(kNm)': nd['Msys'], 'ΔR(kN)': dr,
                    })
                st.dataframe(pd.DataFrame(nd_rows), use_container_width=True, hide_index=True)

        # ── ADS ──
        with vt3:
            show_all_ads = st.checkbox("Mostrar todas ADS", value=False, key="viga_ads_all")
            ads_sorted_v = sorted(sd.get('combo_states_ads', []),
                                  key=lambda c: abs(c['M_total']), reverse=True)
            display_ads = ads_sorted_v if show_all_ads else ads_sorted_v[:10]
            ads_rows = []
            for cs in display_ads:
                bal = cs.get('balance', {})
                row = {'Combo': cs['combo'][:35], 'Grupo': cs.get('group', ''),
                       'M_total': cs['M_total'], 'P_total': cs['P_total'],
                       'ΣR': round(sum(cs.get('reactions_total', [])), 2),
                       'ΣFy': bal.get('sumFy', 0), 'ΣM': bal.get('sumM', 0)}
                for i, fid in enumerate(sd['footings']):
                    rt = cs.get('reactions_total', [])
                    dr = cs.get('delta_R', [])
                    row[f'R_{fid}'] = rt[i] if i < len(rt) else 0
                    row[f'ΔR_{fid}'] = dr[i] if i < len(dr) else 0
                ads_rows.append(row)
            st.dataframe(pd.DataFrame(ads_rows), use_container_width=True, hide_index=True)

        # ── LRFD ──
        with vt4:
            show_all_lrfd = st.checkbox("Mostrar todas LRFD", value=False, key="viga_lrfd_all")
            lrfd_sorted_v = sorted(sd.get('combo_states_lrfd', []),
                key=lambda c: abs(c.get('Vu', 0)) + abs(c.get('Mu_pos', 0)), reverse=True)
            display_lrfd = lrfd_sorted_v if show_all_lrfd else lrfd_sorted_v[:10]
            lrfd_rows = []
            for cs in display_lrfd:
                bal = cs.get('balance', {})
                row = {'Combo': cs['combo'][:35], 'M_total': cs['M_total'],
                       'Vu': cs.get('Vu', 0), 'Mu+': cs.get('Mu_pos', 0), 'Mu-': cs.get('Mu_neg', 0),
                       'ΣR': round(sum(cs.get('reactions_total', [])), 2),
                       'ΣFy': bal.get('sumFy', 0)}
                for i, fid in enumerate(sd['footings']):
                    rt = cs.get('reactions_total', [])
                    dr = cs.get('delta_R', [])
                    row[f'R_{fid}'] = rt[i] if i < len(rt) else 0
                    row[f'ΔR_{fid}'] = dr[i] if i < len(dr) else 0
                lrfd_rows.append(row)
            st.dataframe(pd.DataFrame(lrfd_rows), use_container_width=True, hide_index=True)
            lc1, lc2, lc3 = st.columns(3)
            lc1.metric("φVc", f"{sd['phi_Vc']} kN")
            lc2.metric("As inf", f"{sd['As_inf']}cm² → {sd['As_inf_text']}")
            lc3.metric("As sup", f"{sd['As_sup']}cm² → {sd['As_sup_text']}")

        # ── Resultados ──
        with vt5:
            res_data = {
                'Campo': ['ID', 'Dirección', 'Zapatas', 'Nodos', 'L total', 'ADS ctrl', 'LRFD ctrl',
                          'Mu+', 'Mu-', 'Vu máx', 'φVc', 'sr(%)', 'b viga', 'h viga',
                          'As inf', 'As sup', 's estribo'],
                'Valor': [sd['system_id'], sd['direction'], ', '.join(sd['footings']),
                          sd['num_nodes'], f"{sd['total_length']}m",
                          sd['ads_control'][:30], sd['lrfd_control'][:30],
                          f"{sd['Mu_max_pos']} kN·m", f"{sd['Mu_max_neg']} kN·m",
                          f"{sd['Vu_max']} kN", f"{sd['phi_Vc']} kN",
                          f"{sd['sr_viga']*100:.1f}%", f"{sd['b_viga']*100:.0f}cm",
                          f"{sd['h_viga']*100:.0f}cm",
                          f"{sd['As_inf']}cm² → {sd['As_inf_text']}",
                          f"{sd['As_sup']}cm² → {sd['As_sup_text']}",
                          f"{sd['s_estribo']*100:.0f}cm"]
            }
            st.dataframe(pd.DataFrame(res_data), use_container_width=True, hide_index=True)

            st.divider()
            st.subheader("Estado de zapatas enlazadas")
            zap_rows = []
            for fid in sd['footings']:
                ff = next((f for f in final if f.get('id') == fid), None)
                if not ff: continue
                zap_rows.append({
                    'Zapata': fid, 'B': ff['B'], 'L': ff['L'], 'h': ff['h'],
                    'Pu': ff.get('Pu', 0),
                    'qmax': ff.get('qmax', 0), 'qmax_sys': ff.get('qmax_with_system', ff.get('qmax', 0)),
                    'ratio_sys': ff.get('ratio_with_system', 0),
                    'dP': ff.get('system_dP', 0),
                    'Redim.': '🔴 Sí' if ff.get('needs_resize', False) else '🟢 No',
                })
            st.dataframe(pd.DataFrame(zap_rows), use_container_width=True, hide_index=True)

        # ── Diagramas ──
        with vt6:
            vm = sd.get('vm_diagram')
            if vm and vm.get('stations'):
                fig_vm = go.Figure()
                fig_vm.add_trace(go.Scatter(x=vm['stations'], y=vm['V'], name='V(kN)',
                    line=dict(color='#dc2626', width=2)))
                fig_vm.add_trace(go.Scatter(x=vm['stations'], y=vm['M'], name='M(kN·m)',
                    line=dict(color='#2563eb', width=2)))
                fig_vm.add_hline(y=0, line_dash="dot", line_color="#9ca3af")
                fig_vm.update_layout(title="Diagrama V-M de la viga")
                apply_theme(fig_vm, height=350, equal_axes=False)
                st.plotly_chart(fig_vm, use_container_width=True, key=f"vm_sys_{sd['system_id']}")

                sdg = sd.get('steel_diagram')
                if sdg and sdg.get('stations'):
                    fig_as = go.Figure()
                    fig_as.add_trace(go.Scatter(x=sdg['stations'], y=sdg['As_inf'], name='As inf',
                        fill='tozeroy', line=dict(color='#059669', width=2),
                        fillcolor='rgba(5,150,105,0.15)'))
                    fig_as.add_trace(go.Scatter(x=sdg['stations'],
                        y=[-v for v in sdg['As_sup']], name='As sup',
                        fill='tozeroy', line=dict(color='#d97706', width=2),
                        fillcolor='rgba(217,119,6,0.15)'))
                    fig_as.add_hline(y=0, line_dash="dot", line_color="#9ca3af")
                    fig_as.update_layout(title="Acero requerido [+inf, -sup]")
                    apply_theme(fig_as, height=280, equal_axes=False)
                    st.plotly_chart(fig_as, use_container_width=True, key=f"as_sys_{sd['system_id']}")
            else:
                st.info("Diagramas V-M no disponibles para este sistema.")

        # ── Nota intra-combinada ──
        if sd.get('intra_combined'):
            st.info(f"ℹ️ {sd.get('note', 'Sistema intra-combinado — análisis longitudinal de la zapata.')}")

        # ── Exportar modelo analítico S2K de este sistema ──
        st.divider()
        st.subheader("📤 Exportar modelo analítico de viga")
        _s2k_sys_key = f"s2k_sys_{sd['system_id']}"

        def _gen_s2k_for_system(sys_data, final_footings_list, params_d, sap_units_str,
                                jloads_d, combos_d, lp_d):
            """
            Genera un .$2k 1D completo para el sistema de viga de enlace.
            Todos los nudos se proyectan sobre el eje de la viga (modelo 1D estricto).
            - Cargas en nudos de columna (posición proyectada en eje).
            - Apoyos en centroide de zapata (posición proyectada en eje).
            - Zapata combinada intra-combinada → apoyo FIJO (clamp) para estabilidad.
            - Viga inter-zapatas → apoyo PIN (solo traslaciones).
            - Tabla FRAME SECTION ASSIGNMENTS separada de CONNECTIVITY - FRAME.
            """
            direction  = sys_data.get('direction', 'X')
            sys_id     = sys_data.get('system_id', 'SYS')
            b_v        = float(sys_data.get('b_viga', 0.30))
            h_v        = float(sys_data.get('h_viga', 0.50))
            fc_mpa     = float(params_d.get('fc', 21))
            gc_kn      = float(params_d.get('gamma_c', 24.0))   # kN/m³ peso concreto
            intra_comb = bool(sys_data.get('intra_combined', False))
            sec_name   = f"VE_{int(round(b_v*100))}x{int(round(h_v*100))}"
            mat_name   = f"CONC_{int(round(fc_mpa))}MPa"

            # Aplanar joint strings (pueden ser 'J1,J2' para combined nodes)
            def _flat(s):
                return [j.strip() for j in str(s).split(',') if j.strip()]

            # ── Mapa fid → zapata (para peso propio de zapatas en caso SW) ──────
            f_fmap = {_f.get('id', ''): _f for _f in final_footings_list}

            # ── Mapa joint → datos de columna ────────────────────────────────────
            j2col_map = {}
            for _f in final_footings_list:
                for _c in _f.get('cols', []):
                    j2col_map[str(_c['joint'])] = _c

            # ── Geometría ─────────────────────────────────────────────────────────
            geom       = sys_data.get('geometry', {})
            jnds_per   = sys_data.get('joints', [])   # ['J1', 'J3,J2', ...] — por nodo

            if intra_comb:
                # ── Geometría intra-combinada: reconstruir desde posiciones reales ─
                # analyze_tie_system devuelve solo un punto ficticio en el centroide;
                # para exportar a SAP2000 necesitamos:
                #   • Un nudo de APOYO fijo en el centroide de la zapata.
                #   • Un nudo de CARGA por cada columna en su posición proyectada.
                #   • Viga continua de extremo a extremo (voladizos a ambos lados).
                _fid_ic   = sys_data.get('footings', [''])[0]
                _fobj_ic  = f_fmap.get(_fid_ic, {})
                _cols_ic  = _fobj_ic.get('cols', [])
                _xsup_ic  = (_fobj_ic.get('x_footing', _fobj_ic.get('x', 0))
                             if direction == 'X'
                             else _fobj_ic.get('y_footing', _fobj_ic.get('y', 0)))
                # Joints del sistema (todos los que forman el intra-combinado)
                _all_j_ic = []
                for _js in jnds_per:
                    _all_j_ic.extend(_flat(_js))
                # Coord perp = promedio de columnas involucradas
                _perp_ic = []
                for _j in _all_j_ic:
                    _cc = j2col_map.get(_j)
                    if _cc:
                        _perp_ic.append(_cc['y'] if direction == 'X' else _cc['x'])
                beam_perp = round(sum(_perp_ic) / len(_perp_ic), 5) if _perp_ic else 0.0
                # Construir load_nodes reales: uno por columna con sus joints
                load_nodes = []
                jnds_per_new = []
                for _j in sorted(_all_j_ic,
                                  key=lambda _jj: (j2col_map[_jj]['x'] if direction == 'X'
                                                   else j2col_map[_jj]['y'])
                                  if _jj in j2col_map else 0):
                    _cc = j2col_map.get(_j)
                    if not _cc:
                        continue
                    _xload = _cc['x'] if direction == 'X' else _cc['y']
                    load_nodes.append({
                        'fid': _fid_ic,
                        'x_load': round(_xload, 4),
                        'x_sup':  round(_xsup_ic, 4),
                        'ecc':    round(_xload - _xsup_ic, 4),
                        'coalesced': abs(_xload - _xsup_ic) < 0.01,
                    })
                    jnds_per_new.append(_j)
                jnds_per = jnds_per_new
                if not load_nodes:
                    return None, "No se encontraron columnas para la zapata intra-combinada."
            else:
                load_nodes = geom.get('loads', [])
                if not load_nodes:
                    return None, "Geometría del sistema no disponible (sin nodos de carga)."

            all_j_flat = []
            for _js in jnds_per:
                all_j_flat.extend(_flat(_js) if not intra_comb else [_js])

            if not intra_comb:
                # Coordenada perpendicular del eje para sistemas inter-zapatas
                perp_vals = []
                for _j in all_j_flat:
                    _c = j2col_map.get(_j)
                    if _c:
                        perp_vals.append(_c['y'] if direction == 'X' else _c['x'])
                beam_perp = round(sum(perp_vals) / len(perp_vals), 5) if perp_vals else 0.0

            # ── Registro de nudos 1D (todos sobre el eje de la viga) ─────────
            _pt_map  = {}    # axis_pos → sap_jid
            _pt_list = []    # [(x, y, z, sap_jid)]
            _jcnt    = [1]

            def _get_jnt(axis_pos):
                """Retorna (crea si es nuevo) un nudo SAP en la posición del eje."""
                ap = round(float(axis_pos), 5)
                if ap not in _pt_map:
                    _jid = str(_jcnt[0]); _jcnt[0] += 1
                    _pt_map[ap] = _jid
                    if direction == 'X':
                        _pt_list.append((ap, beam_perp, 0.0, _jid))
                    else:
                        _pt_list.append((beam_perp, ap, 0.0, _jid))
                return _pt_map[ap]

            # Crear nudos de carga y apoyo por nodo
            load_jnt_by_node = [_get_jnt(ln['x_load']) for ln in load_nodes]
            sup_jnt_by_node  = [_get_jnt(ln['x_sup'])  for ln in load_nodes]
            unique_sup_jnts  = list(dict.fromkeys(sup_jnt_by_node))

            # ── Elementos de viga ─────────────────────────────────────────────
            frames_conn = []   # [(jnt_i, jnt_j)]

            if intra_comb:
                # Viga con voladizos: conectar TODOS los nudos en orden de posición
                # (nudos de carga + nudo de apoyo en el centroide).
                # Se conectan tramos consecutivos entre el nudo más a la izquierda
                # y el más a la derecha, incluyendo el apoyo si está entre ellos.
                _all_axes_ic = sorted(_pt_map.keys())
                if len(_all_axes_ic) < 2:
                    return None, "Zapata intra-combinada con un único punto — no se puede generar viga."
                for _i in range(len(_all_axes_ic) - 1):
                    _jA = _get_jnt(_all_axes_ic[_i])
                    _jB = _get_jnt(_all_axes_ic[_i+1])
                    if _jA != _jB:
                        frames_conn.append((_jA, _jB))
            else:
                # Sistema inter-zapatas: tramos entre posiciones de apoyo ordenadas
                sup_axes = sorted(set(round(float(ln['x_sup']), 5) for ln in load_nodes))
                for _i in range(len(sup_axes) - 1):
                    _jA = _get_jnt(sup_axes[_i])
                    _jB = _get_jnt(sup_axes[_i+1])
                    if _jA != _jB:
                        frames_conn.append((_jA, _jB))

                # Brazos excéntricos: nudo-carga → nudo-apoyo (cuando ecc > 0)
                _arm_done = set()
                for _i, ln in enumerate(load_nodes):
                    if not ln.get('coalesced', True) and abs(ln.get('ecc', 0)) > 1e-4:
                        _jld = load_jnt_by_node[_i]
                        _jsp = sup_jnt_by_node[_i]
                        if _jld != _jsp:
                            _pair = tuple(sorted([_jld, _jsp]))
                            if _pair not in _arm_done:
                                frames_conn.append((_jld, _jsp))
                                _arm_done.add(_pair)

            if not frames_conn:
                all_axes = sorted(_pt_map.keys())
                if len(all_axes) >= 2:
                    frames_conn.append((_get_jnt(all_axes[0]), _get_jnt(all_axes[-1])))
                else:
                    return None, "Sistema con un único punto — no se puede generar viga."

            # ── Patrones de carga ─────────────────────────────────────────────
            _pats_raw = [p for p in sys_data.get('patterns_resolved', []) if p != 'PP_VIGA']
            if not _pats_raw:
                for _j in all_j_flat:
                    if _j in jloads_d:
                        _pats_raw = sorted(jloads_d[_j].keys()); break
            _design_type = {
                'D':'Dead','SD':'Superimposed Dead','L':'Live','Lr':'Roof Live',
                'Le':'Live','S':'Snow','W':'Wind','E':'Seismic',
                'Sx':'Seismic','Sy':'Seismic',
                'Wx+':'Wind','Wx-':'Wind','Wy+':'Wind','Wy-':'Wind',
            }
            for _pk, _pv in (lp_d or {}).items():
                if _pk not in _design_type:
                    _t = str(_pv).lower() if isinstance(_pv, str) else ''
                    if   'dead'  in _t: _design_type[_pk] = 'Dead'
                    elif 'live'  in _t: _design_type[_pk] = 'Live'
                    elif 'seism' in _t or 'quake' in _t: _design_type[_pk] = 'Seismic'
                    elif 'wind'  in _t: _design_type[_pk] = 'Wind'
                    else:               _design_type[_pk] = 'Other'
            _pats = _pats_raw or ['D', 'L']

            # ── Helpers ───────────────────────────────────────────────────────
            def _fv(v):
                if isinstance(v, bool): return "Yes" if v else "No"
                if isinstance(v, int):  return str(v)
                if isinstance(v, float):
                    v = 0.0 if abs(v) < 1e-12 else v
                    t = f"{v:.6f}".rstrip('0').rstrip('.')
                    return t if t else "0"
                return str(v)

            _tbl = {}
            def _add(tname, **kwargs):
                _tbl.setdefault(tname, []).append(
                    "   ".join(f"{k}={_fv(v)}" for k, v in kwargs.items()))

            # ── PROGRAM CONTROL ───────────────────────────────────────────────
            _add("PROGRAM CONTROL",
                 ProgramName="SAP2000", Version=26,
                 CurrUnits=f'"{sap_units_str}"',
                 SteelCode='"AISC 360-10"',
                 ConcCode='"ACI 318-08/IBC2009"',
                 RegenHinge="Yes")

            # ── LOAD PATTERN DEFINITIONS ──────────────────────────────────────
            for _p in _pats:
                _add("LOAD PATTERN DEFINITIONS",
                     LoadPat=f'"{_p}"',
                     DesignType=f'"{_design_type.get(_p,"Other")}"',
                     SelfWtMult=0, AutoLoad="None")
            # SW — peso propio de la viga + peso de zapatas (SelfWtMult=1 para
            # que SAP calcule automáticamente el peso propio de los elementos)
            _add("LOAD PATTERN DEFINITIONS",
                 LoadPat='"SW"', DesignType='"Dead"',
                 SelfWtMult=1, AutoLoad="None")

            # ── LOAD CASE DEFINITIONS ─────────────────────────────────────────
            for _p in _pats:
                _add("LOAD CASE DEFINITIONS",
                     Case=f'"{_p}"', Type="LinStatic",
                     InitialCond="Zero",
                     DesignType=f'"{_design_type.get(_p,"Other")}"',
                     RunCase="Yes")
            _add("LOAD CASE DEFINITIONS",
                 Case='"SW"', Type="LinStatic",
                 InitialCond="Zero", DesignType='"Dead"', RunCase="Yes")

            # ── CASE - STATIC 1 - LOAD ASSIGNMENTS ───────────────────────────
            for _p in _pats:
                _add("CASE - STATIC 1 - LOAD ASSIGNMENTS",
                     Case=f'"{_p}"', LoadType='"Load pattern"',
                     LoadName=f'"{_p}"', LoadSF=1)
            _add("CASE - STATIC 1 - LOAD ASSIGNMENTS",
                 Case='"SW"', LoadType='"Load pattern"',
                 LoadName='"SW"', LoadSF=1)

            # ── MATERIAL ──────────────────────────────────────────────────────
            _E = 4700.0 * math.sqrt(fc_mpa)
            _add("MATERIAL PROPERTIES 01 - GENERAL",
                 Material=f'"{mat_name}"', Type="Concrete",
                 SymType="Isotropic", TempDepend="No", Color="Cyan")
            _add("MATERIAL PROPERTIES 02 - BASIC MECHANICAL PROPERTIES",
                 Material=f'"{mat_name}"', E1=_E, U12=0.2,
                 A1=9.9e-6,
                 UnitMass=round(gc_kn / 9.81, 5),
                 UnitWeight=gc_kn)
            _add("MATERIAL PROPERTIES 03B - CONCRETE DATA",
                 Material=f'"{mat_name}"', Fc=fc_mpa*10.1972,
                 LtWtConc="No", FcsFactor=1,
                 SSType="Parametric-Simple", SSHysType="Takeda",
                 StrainAtFc=0.002219, StrainUltimate=0.005, FinalSlope=0)

            # ── FRAME SECTION ─────────────────────────────────────────────────
            _area = b_v * h_v
            _I33  = b_v * h_v**3 / 12
            _I22  = h_v * b_v**3 / 12
            _J    = min(b_v, h_v) * max(b_v, h_v)**3 / 3 * 0.2
            _AS   = max(0.833 * _area, 1e-6)
            _add("FRAME SECTION PROPERTIES 01 - GENERAL",
                 SectionName=f'"{sec_name}"', Material=f'"{mat_name}"',
                 Shape="Rectangular", t3=h_v, t2=b_v,
                 Area=_area, TorsConst=_J, I33=_I33, I22=_I22, I23=0.0,
                 AS2=_AS, AS3=_AS,
                 AMod=1, A2Mod=1, A3Mod=1, JMod=1, I2Mod=1, I3Mod=1, MMod=1, WMod=1)
            _add("FRAME SECTION PROPERTIES 03 - CONCRETE BEAM",
                 SectionName=f'"{sec_name}"',
                 RebarMatL="A706Gr60", RebarMatC="A706Gr60",
                 TopCover=0.04, BotCover=0.04,
                 TopLeftArea=0, TopRghtArea=0, BotLeftArea=0, BotRghtArea=0)

            # ── JOINT COORDINATES (1D — todos en el eje de la viga) ───────────
            for (_x, _y, _z, _jid) in _pt_list:
                _add("JOINT COORDINATES",
                     Joint=_jid, CoordSys="GLOBAL", CoordType="Cartesian",
                     XorR=_x, Y=_y, Z=_z)

            # ── JOINT RESTRAINTS ──────────────────────────────────────────────
            # intra_combined → FIJO completo (clamp): estabilidad con un solo apoyo.
            # inter-zapata (dir X):
            #   R2 = No  → libre en flexión principal (plano XZ, momento M2)
            #   R1 = Yes → fijo torsión/plano perp. (captura momentos M1 de columnas)
            #   R3 = Yes → fijo rotación horizontal (estabilidad)
            # inter-zapata (dir Y):
            #   R1 = No  → libre en flexión principal (plano YZ, momento M1)
            #   R2 = Yes → fijo plano perp. (captura momentos M2 de columnas)
            #   R3 = Yes → fijo rotación horizontal (estabilidad)
            if intra_comb:
                _R1, _R2, _R3 = "Yes", "Yes", "Yes"
            elif direction == 'X':
                _R1, _R2, _R3 = "Yes", "No", "Yes"
            else:  # direction == 'Y'
                _R1, _R2, _R3 = "No", "Yes", "Yes"
            _rest_done = set()
            for _sjid in unique_sup_jnts:
                if _sjid not in _rest_done:
                    _add("JOINT RESTRAINT ASSIGNMENTS",
                         Joint=_sjid, U1="Yes", U2="Yes", U3="Yes",
                         R1=_R1, R2=_R2, R3=_R3)
                    _rest_done.add(_sjid)

            # ── CONNECTIVITY - FRAME (solo conectividad, sin sección aquí) ───
            for _fi_idx, (_jI, _jJ) in enumerate(frames_conn):
                _add("CONNECTIVITY - FRAME",
                     Frame=_fi_idx+1, JointI=_jI, JointJ=_jJ)

            # ── FRAME SECTION ASSIGNMENTS (sección separada) ──────────────────
            for _fi_idx in range(len(frames_conn)):
                _add("FRAME SECTION ASSIGNMENTS",
                     Frame=_fi_idx+1,
                     AutoSelect='"N.A."',
                     AnalSect=f'"{sec_name}"',
                     MatProp='"Default"')

            # ── JOINT LOADS - FORCE (per patrón, por nodo de carga) ───────────
            # Se aplican AMBOS momentos M1 y M2 en cada nudo de carga:
            #   - El momento en la dirección de la viga → flexión principal del modelo
            #   - El momento perpendicular → capturado por el apoyo con R fijo,
            #     permite auditar el efecto transversal sobre la zapata.
            _jld_added = False
            for _pat in _pats:
                for _ni, ln in enumerate(load_nodes):
                    _jld_sap = load_jnt_by_node[_ni]
                    _nj = _flat(jnds_per[_ni] if _ni < len(jnds_per) else '')
                    _Fz = sum(abs(float(jloads_d.get(_j, {}).get(_pat, {}).get('F3', 0)))
                              for _j in _nj)
                    _m2v = sum(float(jloads_d.get(_j, {}).get(_pat, {}).get('M2', 0))
                               for _j in _nj)
                    _m1v = sum(float(jloads_d.get(_j, {}).get(_pat, {}).get('M1', 0))
                               for _j in _nj)
                    if abs(_Fz) > 0.01 or abs(_m1v) > 0.01 or abs(_m2v) > 0.01:
                        _add("JOINT LOADS - FORCE",
                             Joint=_jld_sap, LoadPat=f'"{_pat}"',
                             CoordSys="GLOBAL",
                             F1=0.0, F2=0.0, F3=-_Fz,
                             M1=_m1v, M2=_m2v, M3=0.0)
                        _jld_added = True

            # ── JOINT LOADS - FORCE: peso propio de zapatas (patrón SW) ─────────
            # Cada apoyo recibe la carga puntual del volumen de su zapata.
            # El peso propio de la viga lo calcula SAP con SelfWtMult=1 en SW.
            _fids_sw_done = set()
            for _ni, ln in enumerate(load_nodes):
                _fid_sw = ln.get('fid', '')
                if _fid_sw in _fids_sw_done:
                    continue          # zapata ya procesada (nodos compartidos)
                _fids_sw_done.add(_fid_sw)
                _fobj_sw = f_fmap.get(_fid_sw)
                if _fobj_sw:
                    _Wf = gc_kn * _fobj_sw.get('B', 0) * _fobj_sw.get('L', 0) * _fobj_sw.get('h', 0)
                    if _Wf > 0.01:
                        _jsp_sw = sup_jnt_by_node[_ni]
                        _add("JOINT LOADS - FORCE",
                             Joint=_jsp_sw, LoadPat='"SW"',
                             CoordSys="GLOBAL",
                             F1=0.0, F2=0.0, F3=-_Wf,
                             M1=0.0, M2=0.0, M3=0.0)
                        _jld_added = True

            if not _jld_added:
                _tbl.setdefault("JOINT LOADS - FORCE", []).append(
                    '$ (sin cargas en patrones definidos)')

            # ── COMBINATION DEFINITIONS ───────────────────────────────────────
            _pats_set  = set(_pats)
            _ads_names = []; _lrfd_names = []

            def _add_combo(cname, factors, ctype="Linear Add"):
                # SW se suma con el mismo factor que D en cada combinación
                # (SW representa el peso propio de la viga + volumen de zapatas)
                _ext = dict(factors)
                if 'D' in _ext:
                    _ext['SW'] = _ext['D']
                _items = [(p, sf) for p, sf in _ext.items()
                          if p in _pats_set or p == 'SW']
                if not _items: return False
                for _k, (_p, _sf) in enumerate(_items):
                    if _k == 0:
                        _add("COMBINATION DEFINITIONS",
                             ComboName=f'"{cname}"', ComboType=f'"{ctype}"',
                             AutoDesign="No", CaseName=f'"{_p}"', ScaleFactor=_sf)
                    else:
                        _add("COMBINATION DEFINITIONS",
                             ComboName=f'"{cname}"', CaseName=f'"{_p}"',
                             ScaleFactor=_sf)
                return True

            def _add_envelope(envname, members):
                for _k, _m in enumerate(members):
                    if _k == 0:
                        _add("COMBINATION DEFINITIONS",
                             ComboName=f'"{envname}"', ComboType='"Envelope"',
                             AutoDesign="No", CaseName=f'"{_m}"', ScaleFactor=1)
                    else:
                        _add("COMBINATION DEFINITIONS",
                             ComboName=f'"{envname}"', CaseName=f'"{_m}"',
                             ScaleFactor=1)

            if combos_d:
                for _c in combos_d.get('ADS', []):
                    if _add_combo(_c['name'], _c['factors']): _ads_names.append(_c['name'])
                for _c in combos_d.get('LRFD', []):
                    if _add_combo(_c['name'], _c['factors']): _lrfd_names.append(_c['name'])
                if _ads_names:  _add_envelope("ENV_ASD",  _ads_names)
                if _lrfd_names: _add_envelope("ENV_LRFD", _lrfd_names)

            # ── Ensamblar texto final ─────────────────────────────────────────
            _out = [f'$ SAP2000 .$2k — Viga de enlace: {sys_id}',
                    '$ Generado por Cimentaciones App (NSR-10)',
                    f'$ Dirección: {direction}  |  Intra-combinada: {"Sí (apoyo fijo)" if intra_comb else "No (apoyo pin)"}',
                    f'$ Apoyos: U1=U2=U3=Yes  R1={_R1}  R2={_R2}  R3={_R3}',
                    f'$ SW: SelfWtMult=1 (peso viga auto) + pesos zapata en apoyos.',
                    f'$ Momentos M1 y M2 aplicados en todos los nudos de carga.',
                    '']
            _order = [
                "PROGRAM CONTROL",
                "LOAD PATTERN DEFINITIONS",
                "LOAD CASE DEFINITIONS",
                "CASE - STATIC 1 - LOAD ASSIGNMENTS",
                "MATERIAL PROPERTIES 01 - GENERAL",
                "MATERIAL PROPERTIES 02 - BASIC MECHANICAL PROPERTIES",
                "MATERIAL PROPERTIES 03B - CONCRETE DATA",
                "FRAME SECTION PROPERTIES 01 - GENERAL",
                "FRAME SECTION PROPERTIES 03 - CONCRETE BEAM",
                "JOINT COORDINATES",
                "JOINT RESTRAINT ASSIGNMENTS",
                "CONNECTIVITY - FRAME",
                "FRAME SECTION ASSIGNMENTS",
                "JOINT LOADS - FORCE",
                "COMBINATION DEFINITIONS",
            ]
            for _tname in _order:
                if _tname not in _tbl or not _tbl[_tname]: continue
                _out.append(f'TABLE:  "{_tname}"')
                for _row in _tbl[_tname]:
                    _out.append(f'   {_row}')
                _out.append('')
            _out.append('END TABLE DATA')
            return '\n'.join(_out), None

        _btn_s2k_sys = st.button(
            f"🏗️ Generar S2K del sistema **{sd['system_id']}**",
            key=f"btn_gen_s2k_{sd['system_id']}", use_container_width=True)
        if _btn_s2k_sys:
            _jloads_sys  = results.get('export_jloads', md.get('_jloads', {}))
            _combos_sys  = __import__('parser').gen_combos(md.get('_cl', {}), R=params.get('R', 1.0))
            _lp_sys      = md.get('_lp', {})
            _s2k_txt, _s2k_err = _gen_s2k_for_system(
                sd, final, params, sap_units,
                _jloads_sys, _combos_sys, _lp_sys)
            if _s2k_err:
                st.error(_s2k_err)
            else:
                st.session_state[_s2k_sys_key] = _s2k_txt
                st.success("Modelo S2K generado.")

        if st.session_state.get(_s2k_sys_key):
            st.download_button(
                f"📥 Descargar {sd['system_id']}.$2k",
                data=st.session_state[_s2k_sys_key],
                file_name=f"{sd['system_id']}.$2k",
                mime="text/plain",
                key=f"dl_s2k_{sd['system_id']}",
                use_container_width=True,
            )

# ════════════════════════════════════════════════
# TAB: EXPORTAR
# ════════════════════════════════════════════════
with tab_export:
    ec1, ec2 = st.columns(2)
    with ec1:
        st.download_button("📥 results.json",
            json.dumps({
                'final_footings': [{k: v for k, v in f.items()
                    if k not in ('ads_audit', 'lrfd_audit', 'column_forces')} for f in final],
                'total_area': results['total_area'],
                'total_volume': results['total_volume']
            }, indent=2, default=str),
            file_name="results.json", mime="application/json")
    with ec2:
        csv_rows = []
        for f in final:
            rx2 = propose_rebar(f['Asx'], f['B'], rec)
            ry2 = propose_rebar(f['Asy'], f['L'], rec)
            csv_rows.append({
                'ID': f['id'], 'Tipo': f.get('type', '?'), 'B': f['B'], 'L': f['L'], 'h': f['h'],
                'qmax': f['qmax'], 'Pu': f['Pu'],
                'P%': round(f['pr']*100, 1), 'V%': round(f['sr']*100, 1),
                'Asx': f['Asx'], 'dx': rx2['text'], 'Asy': f['Asy'], 'dy': ry2['text'],
                'Estado': f['st']
            })
        st.download_button("📥 resumen.csv",
            pd.DataFrame(csv_rows).to_csv(index=False),
            file_name="resumen.csv", mime="text/csv")

    # ── EXPORTACIÓN SAP2000 / ETABS ──
    st.divider()
    st.subheader("🏗️ Exportación del modelo de cimentación")

    _base_model_name = uploaded.name.rsplit(".", 1)[0] + "_CIMENTACION"
    export_cfg = {
        "model_name": _base_model_name,
        "units": sap_units,
        "pedestal_h": pedestal_h_exp,
        "z_top": z_top_exp,
        "k_subgrade": k_subgrade_exp,
        "alpha_xy": alpha_xy_exp,
        "fc_mpa": fc,
        "concrete_material_name": f"CONC_{int(round(fc))}MPa",
        "tie_b": tie_b_exp,
        "tie_h": tie_h_exp,
        "shell_section_prefix": "ZAP_",
        "project_info": {
            "Company Name": "SmartCouplers MG SAS",
            "Client Name": "Sin Nombre",
            "Project Name": "Sin Nombre",
            "Project Number": "001",
            "Model Name": _base_model_name,
            "Model Description": "Modelo creado a partir de una aplicacion web creada con IA",
            "Revision Number": "001",
            "Frame Type": "Sistema de Cimentaciones Superficiales",
            "Engineer": "Creado Por - Sergio Medina",
            "Checker": "Definir Nombre de Revisor",
            "Supervisor": "Definir Nombre de Supervisor",
            "Issue Code": "Version 01",
            "Design Code": "Version 01",
        }
    }

    sx1, sx2 = st.columns(2)

    with sx1:
        st.markdown("**SAP2000 (.$2k)**")
        if st.button("🧱 Generar .$2k SAP2000", type="primary", use_container_width=True, key="btn_generar_s2k_export_tab"):
            try:
                md_export = dict(md)
                md_export["_jloads"] = results.get("export_jloads", md.get("_jloads", {}))
                s2k_text = export_foundation_s2k(
                    model_data=md_export,
                    results=results,
                    params=params,
                    export_cfg=export_cfg
                )
                if 'TABLE:  "PROGRAM CONTROL"' not in s2k_text:
                    raise ValueError("El .$2k generado no contiene la tabla PROGRAM CONTROL")
                st.session_state["sap_s2k_text"] = s2k_text
                st.success("Archivo .$2k generado correctamente.")
            except Exception as e:
                st.error(f"Error al generar SAP2000: {e}")

        if "sap_s2k_text" in st.session_state:
            st.download_button(
                "📥 Descargar .$2k",
                data=st.session_state["sap_s2k_text"],
                file_name=f"{_base_model_name}.$2k",
                mime="text/plain",
                use_container_width=True,
                key="btn_descargar_s2k_export_tab",
            )

    with sx2:
        st.markdown("**ETABS (.e2k)**")
        st.info("🔧 Exportación ETABS temporalmente suspendida — en revisión.")

    st.caption("Los archivos se generan únicamente desde esta pestaña.")

    # ── EXCEL COMPLETO ──
    st.divider()
    if st.button("🔨 Generar informe Excel", type="primary"):
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
            from openpyxl.drawing.image import Image as XLImage
        except ImportError:
            st.error("pip install openpyxl"); st.stop()
        import os as _os
        from datetime import datetime as _dt

        wb = openpyxl.Workbook()
        hf = Font(bold=True, color='FFFFFF', size=10)
        hfill = PatternFill(start_color='2563EB', end_color='2563EB', fill_type='solid')
        okf = PatternFill(start_color='D1FAE5', fill_type='solid')
        wf = PatternFill(start_color='FEF3C7', fill_type='solid')
        ff_fill = PatternFill(start_color='FEE2E2', fill_type='solid')
        bd = Border(left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin'))

        def wh(ws, row, headers):
            for c, h in enumerate(headers, 1):
                cell = ws.cell(row=row, column=c, value=h)
                cell.font = hf; cell.fill = hfill
                cell.alignment = Alignment(horizontal='center'); cell.border = bd

        # ── 0. PORTADA ──
        ws0 = wb.active; ws0.title = "Portada"

        # Anchos de columna: A(margen) B(nombre hoja) C-D(descripción) E-H(variables) I(margen)
        for ci, cw in zip(range(1, 10), [2, 16, 16, 16, 16, 16, 16, 16, 2]):
            ws0.column_dimensions[get_column_letter(ci)].width = cw

        # --- Logo: crop 25% superior e inferior → zona central 50% → display 4"×2" ---
        _logo_path = _os.path.join(_os.path.dirname(_os.path.abspath(__file__)), 'assets', 'logo.png')
        if _os.path.exists(_logo_path):
            try:
                from PIL import Image as _PILImg
                _pil = _PILImg.open(_logo_path)
                _pw_orig, _ph_orig = _pil.size
                _pil_crop = _pil.crop((0, _ph_orig // 4, _pw_orig, 3 * _ph_orig // 4))
                _logo_buf = io.BytesIO()
                _pil_crop.save(_logo_buf, format='PNG')
                _logo_buf.seek(0)
                _img = XLImage(_logo_buf)
            except Exception:
                _img = XLImage(_logo_path)
            _img.width = 384   # 4" a 96 dpi
            _img.height = 192  # 2" a 96 dpi
            ws0.add_image(_img, 'B2')
        for _r in range(2, 6):
            ws0.row_dimensions[_r].height = 28   # 4 filas × 28pt
        ws0.row_dimensions[6].height = 14

        # --- Título (a la derecha del logo) ---
        ws0.merge_cells('E2:H5')
        _c = ws0['E2']
        _c.value = "DISEÑO DE CIMENTACIONES SUPERFICIALES"
        _c.font = Font(bold=True, size=28, color='1E3A5F')
        _c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        ws0.merge_cells('E6:H6')
        ws0.row_dimensions[6].height = 14
        _c = ws0['E6']
        _c.value = f"Informe generado: {_dt.now().strftime('%d/%m/%Y  %H:%M')}"
        _c.font = Font(size=9, color='888888')
        _c.alignment = Alignment(horizontal='right', vertical='center')

        # --- Separador azul ---
        ws0.row_dimensions[7].height = 5
        for _ci in range(2, 9):
            ws0.cell(row=7, column=_ci).fill = PatternFill(start_color='2563EB', end_color='2563EB', fill_type='solid')

        # helper: escribe + fusiona B:H en una fila
        def _pw(row, value, font_kw=None, fill_color=None, h=16, wrap=False, align='left'):
            ws0.merge_cells(start_row=row, end_row=row, start_column=2, end_column=8)
            ws0.row_dimensions[row].height = h
            _cell = ws0.cell(row=row, column=2, value=value)
            _cell.font = Font(**(font_kw or {}))
            if fill_color:
                _cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')
            _cell.alignment = Alignment(horizontal=align, vertical='center', wrap_text=wrap)
            return _cell

        _r = 8   # spacer
        ws0.row_dimensions[_r].height = 6

        # ── SMART COUPLERS MG ──
        _r = 9
        _pw(_r, "  SMART COUPLERS MG",
            font_kw=dict(bold=True, size=11, color='FFFFFF'), fill_color='1E3A5F', h=20)
        for _sc_line in [
            ("Smart Couplers MG es una empresa colombiana dedicada a la implementación de métodos y técnicas "
             "industrializadas en la construcción de estructuras, principalmente de concreto reforzado."),
            ("Distribuimos y asesoramos en la implementación de empalmes mecánicos en sus proyectos. "
             "Esta aplicación fue desarrollada como parte de nuestras iniciativas de innovación tecnológica."),
            ("Sitio web: www.scmgsas.com   |   Tel: +57 323 2849503   |   gerencia@scmgsas.com"),
        ]:
            _r += 1
            _is_contact = _sc_line.startswith("Sitio web")
            _pw(_r, _sc_line,
                font_kw=dict(size=10, bold=_is_contact, color='1E3A5F' if _is_contact else '1F2937'),
                fill_color='EFF6FF', h=30, wrap=True)

        # ── DESCRIPCIÓN ──
        _r += 2
        _pw(_r, "  DESCRIPCIÓN DE LA APLICACIÓN",
            font_kw=dict(bold=True, size=11, color='FFFFFF'), fill_color='2563EB', h=20)

        _desc_lines = [
            "Esta herramienta realiza el predimensionamiento y diseño de cimentaciones superficiales aisladas y combinadas.",
            "Procesa modelos estructurales, clasifica zapatas según su ubicación (concéntrica, medianera, esquinera),",
            "dimensiona geométricamente por presiones admisibles (ASD) y verifica el diseño por resistencia última (LRFD-ACI 318).",
            "Incluye diseño de vigas de cimentación (sistemas de enlace) con análisis de momentos, cortantes y refuerzo.",
            "Desarrollada con el modelo de Inteligencia Artificial  Claude Opus 4.6  de Anthropic.",
        ]
        for _i, _line in enumerate(_desc_lines):
            _r += 1
            _is_ai = _line.startswith("Desarrollada")
            _pw(_r, _line,
                font_kw=dict(size=10, bold=_is_ai),
                fill_color='EFF6FF' if _is_ai else 'F9FAFB',
                h=16, wrap=True)

        # ── BASE TEÓRICA ──
        _r += 2
        _pw(_r, "  BASE TEÓRICA",
            font_kw=dict(bold=True, size=11, color='FFFFFF'), fill_color='2563EB', h=20)
        for _book in [
            "• Bowles, J.E. — Foundation Analysis and Design, 5th Ed. — McGraw-Hill",
            "• McCormac, J.C. — Diseño de Concreto Reforzado, 8th Ed. — Alfaomega / Wiley",
        ]:
            _r += 1
            _pw(_r, _book, font_kw=dict(size=10, italic=True), fill_color='F9FAFB', h=16)

        # ── ADVERTENCIAS ──
        _r += 2
        _pw(_r, "  ⚠   ADVERTENCIAS IMPORTANTES — LEER ANTES DE USAR",
            font_kw=dict(bold=True, size=11, color='FFFFFF'), fill_color='B91C1C', h=22)

        _warnings = [
            ("Esta aplicación NO ha sido auditada ni probada exhaustivamente. Sus resultados pueden contener errores "
             "no detectados. SIEMPRE verifique los resultados con un ingeniero estructural calificado antes de "
             "cualquier uso.", True),
            ("La herramienta fue generada con Inteligencia Artificial (IA). Los modelos de IA pueden producir "
             "resultados plausibles pero incorrectos (fenómeno conocido como 'alucinaciones'). No se garantiza "
             "la exactitud, completitud ni consistencia de ningún cálculo.", True),
            ("La IA que generó esta aplicación no posee licencia de ingeniería ni asume responsabilidad "
             "profesional. No puede reemplazar el criterio y la responsabilidad de un profesional habilitado.", False),
            ("USO EXCLUSIVAMENTE DIDÁCTICO Y EDUCATIVO. Esta aplicación NO es válida para uso profesional, "
             "proyectos de construcción reales, trámites de permiso ni toma de decisiones estructurales.", True),
            ("El uso de esta herramienta con fines profesionales o en proyectos reales queda bajo la entera "
             "responsabilidad del usuario. Los autores no asumen ninguna responsabilidad por daños derivados.", False),
        ]
        for _warn, _crit in _warnings:
            _r += 1
            _pw(_r, _warn,
                font_kw=dict(size=10, bold=_crit, color='7F1D1D' if _crit else '1F2937'),
                fill_color='FEE2E2' if _crit else 'FEF3C7',
                h=36, wrap=True)

        # ── CONTENIDO DEL INFORME ──
        _r += 2
        _pw(_r, "  CONTENIDO DEL INFORME — GUÍA DE HOJAS",
            font_kw=dict(bold=True, size=11, color='FFFFFF'), fill_color='2563EB', h=20)

        # Encabezados de la tabla guía
        _r += 1
        ws0.row_dimensions[_r].height = 18
        for _sc, _ec, _htitle in [(2, 2, "Hoja"), (3, 4, "Descripción"), (5, 8, "Variables principales")]:
            ws0.merge_cells(start_row=_r, end_row=_r, start_column=_sc, end_column=_ec)
            _c = ws0.cell(row=_r, column=_sc, value=_htitle)
            _c.font = Font(bold=True, size=10, color='FFFFFF')
            _c.fill = PatternFill(start_color='374151', end_color='374151', fill_type='solid')
            _c.alignment = Alignment(horizontal='center', vertical='center')
            _c.border = bd

        _sheet_guide = [
            ("Resumen",
             "Dimensiones finales de cada zapata y resumen ejecutivo de diseño. "
             "Los valores marcados con * corresponden a la combinación LRFD crítica (mayor ratio de punzonamiento).",
             "ID — identificador; Tipo — simple/combined; Esquema — tipo de zapata; "
             "Loc — ubicación (concéntrica, medianera, esquinera); B × L × h — dimensiones (m); "
             "Área (m²); Vol (m³); qmax / qmin — presiones extremas del suelo (kPa); "
             "FSv — factor de seguridad al volcamiento; FSd — factor de seguridad al deslizamiento; "
             "Pu — carga axial última máxima (kN); "
             "Mux* / Muy* — momentos últimos de la combo crítica (kN·m); "
             "ex* / ey* — excentricidades = |Muy|/Pu y |Mux|/Pu de la combo crítica (m); "
             "Vu_x / Vu_y — cortante último 1-vía (viga amplia) por dirección, combo crítica (kN); "
             "Vx% / Vy% — demanda/capacidad a cortante 1-vía por dirección (%); "
             "Vu_pun — fuerza última de punzonamiento = Pu·(1 − Ap/A), combo crítica (kN); "
             "Pun% — ratio de demanda/capacidad a punzonamiento = vu_max/φvc (%); "
             "Asx / Asy — acero requerido por dirección (cm²); Ref X / Ref Y — varilla propuesta; "
             "Estado — resultado del diseño (OK, REVISION, NO_CUMPLE)"),
            ("ASD",
             "Verificación por presiones admisibles del suelo (Allowable Stress Design). "
             "Una fila por combinación de carga por cada zapata.",
             "Combo — nombre de la combinación de carga; Grupo q1/q2/q3 — tipo de carga "
             "(permanente, temporal, sísmica); qadm — presión admisible del suelo (kPa); "
             "P+Wp — carga vertical total incluyendo peso de zapata (kN); "
             "Mx / My — momentos en la base de la zapata (kN·m); "
             "Vx / Vy — cortantes horizontales (kN); ex / ey — excentricidades resultantes (m); "
             "Contacto — estado de contacto suelo-zapata (full / parcial); "
             "q1, q2, q3, q4 — presiones en las cuatro esquinas (kPa); "
             "qmax / qmin — presión máxima y mínima (kPa); "
             "Ratio = qmax / qadm (verde ≤ 0.85 | amarillo ≤ 1.05 | rojo > 1.05); "
             "FSv — factor de seguridad al volcamiento; FSd — factor de seguridad al deslizamiento"),
            ("LRFD",
             "Diseño por resistencia última (Load & Resistance Factor Design — ACI 318). "
             "Una fila por combinación de carga, ordenadas de mayor a menor ratio de punzonamiento.",
             "Combo — combinación de carga última; Pu — carga axial última (kN); "
             "Mux / Muy — momentos últimos por dirección (kN·m); "
             "ex = |Muy|/Pu — excentricidad en X (m); ey = |Mux|/Pu — excentricidad en Y (m); "
             "── PUNZONAMIENTO (ACI 318 §22.6) ── "
             "Vu_pun = Pu·(1 − Ap/A) — fuerza última de punzonamiento (kN); "
             "Ap = área del perímetro crítico bo·d (d/2 desde la cara de columna); "
             "vu_pun — tensión de corte unitaria combinada = vu_d + γvMu/(J·c) (kPa); "
             "φvc_pun — resistencia reducida del concreto a punzonamiento (kPa); "
             "Pun% = vu_pun/φvc_pun × 100 (%); "
             "── CORTANTE 1-VÍA (viga amplia) ── "
             "Vu_x / Vu_y — cortante último en sección crítica (a distancia d de cara de columna) por dirección (kN); "
             "φVn_x / φVn_y — resistencia reducida φ·0.17·√(f'c)·b·d por dirección (kN); "
             "Vx% / Vy% — ratios de corte por dirección (%); Vcorte% — máximo entre Vx% y Vy% (%); "
             "Mu_x / Mu_y — momentos de diseño a flexión en sección crítica (kN·m); "
             "As_x / As_y — acero de refuerzo requerido por dirección (cm²)"),
            ("Cantidades",
             "Metrado de materiales de la cimentación (material take-off). "
             "Una fila por zapata.",
             "ID — identificador de zapata; Tipo — simple / combined; "
             "B, L, h — dimensiones de la zapata (m); "
             "Vol(m³) — volumen de concreto de la zapata; "
             "Concreto(kN) — peso propio del concreto (γc × Vol); "
             "Asx / Asy — área de acero de refuerzo requerida por dirección (cm²); "
             "Ref X / Ref Y — especificación de varilla propuesta (cantidad × diámetro)"),
            ("Sistemas\nEnlace",
             "Diseño de vigas de cimentación que conectan las zapatas entre sí. "
             "Una fila por sistema de enlace.",
             "Sistema — identificador del sistema (S01, S02...); Dir — dirección X o Y; "
             "Zapatas — IDs de zapatas conectadas; Nodos — cantidad de apoyos del sistema; "
             "L total(m) — longitud total de la viga de enlace; "
             "Mu+(kNm) — momento flector positivo máximo (zona de tracción inferior); "
             "Mu-(kNm) — momento flector negativo máximo (zona de tracción superior); "
             "Vu(kN) — cortante máximo; b × h (cm) — sección transversal de la viga; "
             "As inf(cm²) — acero longitudinal inferior; Ref inf — varilla inferior propuesta; "
             "As sup(cm²) — acero longitudinal superior; Ref sup — varilla superior propuesta; "
             "V% — ratio de demanda / capacidad a cortante (%); "
             "s_est(cm) — espaciado de estribos; Estado — ok / REVISAR_SECCION"),
        ]
        for _sname, _sdesc, _svars in _sheet_guide:
            _r += 1
            ws0.row_dimensions[_r].height = 65
            # Columna nombre hoja
            ws0.cell(row=_r, column=2, value=_sname).font = Font(bold=True, size=10, color='1E3A5F')
            ws0.cell(row=_r, column=2).fill = PatternFill(start_color='DBEAFE', end_color='DBEAFE', fill_type='solid')
            ws0.cell(row=_r, column=2).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws0.cell(row=_r, column=2).border = bd
            # Columna descripción (C:D fusionadas)
            ws0.merge_cells(start_row=_r, end_row=_r, start_column=3, end_column=4)
            _c = ws0.cell(row=_r, column=3, value=_sdesc)
            _c.font = Font(size=9); _c.border = bd
            _c.fill = PatternFill(start_color='F9FAFB', end_color='F9FAFB', fill_type='solid')
            _c.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            # Columna variables (E:H fusionadas)
            ws0.merge_cells(start_row=_r, end_row=_r, start_column=5, end_column=8)
            _c = ws0.cell(row=_r, column=5, value=_svars)
            _c.font = Font(size=8); _c.border = bd
            _c.fill = PatternFill(start_color='F9FAFB', end_color='F9FAFB', fill_type='solid')
            _c.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

        # ── 1. Resumen ──
        ws = wb.create_sheet("Resumen")
        ws['A1'] = "PREDIMENSIONAMIENTO DE CIMENTACIONES"; ws['A1'].font = Font(bold=True, size=14)
        ws['A2'] = f"R={R} | f'c={fc}MPa | qadm: {qadm_1}/{qadm_2}/{qadm_3} kPa"
        headers = ['ID', 'Tipo', 'Esquema', 'Loc', 'B', 'L', 'h', 'Área', 'Vol',
                   'qmax', 'qmin', 'FSv', 'FSd',
                   'Pu (kN)', 'Mux* (kN·m)', 'Muy* (kN·m)', 'ex* (m)', 'ey* (m)',
                   'Vu_x (kN)', 'Vx%', 'Vu_y (kN)', 'Vy%',
                   'Vu_pun (kN)', 'Pun%',
                   'Asx', 'Ref X', 'Asy', 'Ref Y', 'Estado']
        wh(ws, 4, headers)
        for i, f in enumerate(final, 5):
            rx3 = propose_rebar(f['Asx'], f['B'], rec)
            ry3 = propose_rebar(f['Asy'], f['L'], rec)
            loc = f.get('classification', {}).get('location', '') if f.get('type') != 'combined' else f.get('scheme', '')
            vol = f['B'] * f['L'] * f['h']
            fsv = f.get('fs_volc_min', 999); fsd = f.get('fs_desl_min', 999)
            _lc3 = max(f.get('lrfd_audit', [{}]), key=lambda a: a.get('punch_ratio', 0), default={})
            vals = [f['id'], f.get('type', '?'), f.get('scheme', ''), loc[:12],
                    f['B'], f['L'], f['h'], f['A'], round(vol, 3),
                    f['qmax'], f['qmin'],
                    round(fsv, 1) if fsv < 900 else '∞', round(fsd, 1) if fsd < 900 else '∞',
                    f['Pu'],
                    _lc3.get('Mux', 0), _lc3.get('Muy', 0),
                    _lc3.get('ex', 0), _lc3.get('ey', 0),
                    _lc3.get('Vu_x', 0), round(_lc3.get('sr_x', 0)*100, 1),
                    _lc3.get('Vu_y', 0), round(_lc3.get('sr_y', 0)*100, 1),
                    _lc3.get('Vu_punch', 0), round(_lc3.get('punch_ratio', 0)*100, 1),
                    f['Asx'], rx3['text'], f['Asy'], ry3['text'], f['st']]
            for c, v in enumerate(vals, 1):
                cell = ws.cell(row=i, column=c, value=v); cell.border = bd
                if c == len(vals):
                    cell.fill = okf if 'OK' in str(v) else (wf if 'REVISION' in str(v) else ff_fill)
        for c in range(1, len(headers)+1):
            ws.column_dimensions[get_column_letter(c)].width = 14

        # 2. ADS
        ws2 = wb.create_sheet("ADS"); ws2['A1'] = "VERIFICACIÓN ADS"; ws2['A1'].font = Font(bold=True, size=12)
        row_n = 3
        for f in final:
            ws2.cell(row=row_n, column=1, value=f"═══ {f['id']} — {f['B']}×{f['L']}m ═══").font = Font(bold=True)
            row_n += 1
            ah = ['Combo', 'Grupo', 'qadm', 'P+Wp', 'Mx', 'My', 'Vx', 'Vy', 'ex', 'ey', 'Contacto',
                'q1', 'q2', 'q3', 'q4', 'qmax', 'qmin', 'Ratio', 'FSv', 'FSd']
            wh(ws2, row_n, ah); row_n += 1
            for a in sorted(f.get('ads_audit', []), key=lambda x: x['ratio'], reverse=True):
                fsv = a.get('fs_volc_min', 999); fsd = a.get('fs_desl_min', 999)
                vals = [a['combo'], a['group'], a['qadm'], a['P_total'], a['Mx'], a['My'],
                    a.get('Vx', 0), a.get('Vy', 0),
                    a.get('ex', 0), a.get('ey', 0), a.get('contact', '?'),
                    a.get('q1', 0), a.get('q2', 0), a.get('q3', 0), a.get('q4', 0),
                    a['qmax'], a['qmin'], a['ratio'],
                    round(fsv, 1) if fsv < 900 else 999, round(fsd, 1) if fsd < 900 else 999]
                for c, v in enumerate(vals, 1):
                    cell = ws2.cell(row=row_n, column=c, value=v); cell.border = bd
                    if c == len(vals):
                        cell.fill = ff_fill if v > 1.05 else (wf if v > 0.85 else okf)
                row_n += 1
            # Cargas por columna (solo combinadas)
            _cal = f.get('col_ads_loads', [])
            if _cal:
                _ch = ws2.cell(row=row_n, column=1, value=f"  ↳ Cargas por columna — {f['id']}")
                _ch.font = Font(italic=True, color='374151'); row_n += 1
                _cal_h = ['Columna', 'Combo', 'Grupo', 'P (kN)', 'Mx (kN·m)', 'My (kN·m)', 'Vx (kN)', 'Vy (kN)']
                wh(ws2, row_n, _cal_h); row_n += 1
                for _a in sorted(_cal, key=lambda x: (x['Columna'], x['P (kN)'])):
                    _av = [_a['Columna'], _a['Combo'], _a['Grupo'],
                           _a['P (kN)'], _a['Mx (kN·m)'], _a['My (kN·m)'],
                           _a['Vx (kN)'], _a['Vy (kN)']]
                    for c, v in enumerate(_av, 1):
                        ws2.cell(row=row_n, column=c, value=v).border = bd
                    row_n += 1
            row_n += 1

        # 3. LRFD
        ws3 = wb.create_sheet("LRFD"); ws3['A1'] = "DISEÑO LRFD"; ws3['A1'].font = Font(bold=True, size=12)
        row_n = 3
        for f in final:
            ws3.cell(row=row_n, column=1, value=f"═══ {f['id']} — d={f.get('d', 0)}m ═══").font = Font(bold=True)
            row_n += 1
            lh = ['Combo', 'Pu (kN)', 'Mux (kN·m)', 'Muy (kN·m)', 'ex (m)', 'ey (m)',
                  # ── Punzonamiento ACI 318 §22.6 ──
                  'Vu_pun (kN)', 'vu_pun (kPa)', 'φvc_pun (kPa)', 'Pun%',
                  # ── Cortante 1-vía X ──
                  'Vu_x (kN)', 'φVn_x (kN)', 'Vx%',
                  # ── Cortante 1-vía Y ──
                  'Vu_y (kN)', 'φVn_y (kN)', 'Vy%',
                  'Vcorte%',
                  'Mu_x (kN·m)', 'Mu_y (kN·m)', 'As_x (cm²)', 'As_y (cm²)']
            wh(ws3, row_n, lh); row_n += 1
            for a in sorted(f.get('lrfd_audit', []), key=lambda x: x.get('punch_ratio', 0), reverse=True):
                vals = [a['combo'], a['Pu'], a['Mux'], a['Muy'],
                    a.get('ex', 0), a.get('ey', 0),
                    a.get('Vu_punch', 0), a.get('vu_max', 0), a.get('phi_vc', 0),
                    round(a.get('punch_ratio', 0)*100, 1),
                    a.get('Vu_x', 0), a.get('phi_Vn_x', 0), round(a.get('sr_x', 0)*100, 1),
                    a.get('Vu_y', 0), a.get('phi_Vn_y', 0), round(a.get('sr_y', 0)*100, 1),
                    round(a.get('shear_ratio', 0)*100, 1),
                    a.get('Mu_x', 0), a.get('Mu_y', 0), a.get('As_x', 0), a.get('As_y', 0)]
                for c, v in enumerate(vals, 1):
                    cell = ws3.cell(row=row_n, column=c, value=v); cell.border = bd
                row_n += 1
            # Punzonamiento por columna (solo combinadas)
            _cpa = f.get('col_punch_audit', [])
            if _cpa:
                _ph = ws3.cell(row=row_n, column=1,
                               value=f"  ↳ Punzonamiento por columna — {f['id']}  (ACI 318 §22.6)")
                _ph.font = Font(italic=True, color='374151'); row_n += 1
                _cpa_h = ['Columna', 'Combo', 'bx (m)', 'by (m)',
                           'Pu (kN)', 'Mux (kN·m)', 'Muy (kN·m)', 'ex (m)', 'ey (m)',
                           'Vu_pun (kN)', 'vu_pun (kPa)', 'φvc_pun (kPa)', 'Pun%']
                wh(ws3, row_n, _cpa_h); row_n += 1
                for _a in sorted(_cpa, key=lambda x: (x['Columna'], -x['Pun%'])):
                    _av = [_a['Columna'], _a['Combo'], _a['bx (m)'], _a['by (m)'],
                           _a['Pu (kN)'], _a['Mux (kN·m)'], _a['Muy (kN·m)'],
                           _a['ex (m)'], _a['ey (m)'],
                           _a['Vu_pun (kN)'], _a['vu_pun (kPa)'], _a['φvc_pun (kPa)'], _a['Pun%']]
                    for ci, v in enumerate(_av, 1):
                        _cell = ws3.cell(row=row_n, column=ci, value=v); _cell.border = bd
                        if ci == len(_av):   # colorear Pun%
                            if isinstance(v, (int, float)):
                                _cell.fill = ff_fill if v > 100 else (wf if v > 75 else okf)
                    row_n += 1
            row_n += 1

        # 4. Cantidades
        ws4 = wb.create_sheet("Cantidades"); ws4['A1'] = "CANTIDADES DE OBRA"
        ws4['A1'].font = Font(bold=True, size=12)
        qh = ['ID', 'Tipo', 'B', 'L', 'h', 'Vol(m3)', 'Concreto(kN)', 'Asx', 'Ref X', 'Asy', 'Ref Y']
        wh(ws4, 3, qh)
        for i, f in enumerate(final, 4):
            rx4 = propose_rebar(f['Asx'], f['B'], rec)
            ry4 = propose_rebar(f['Asy'], f['L'], rec)
            vol = f['B'] * f['L'] * f['h']
            vals = [f['id'], f.get('type', '?'), f['B'], f['L'], f['h'],
                    round(vol, 3), round(vol*gc, 0),
                    f['Asx'], rx4['text'], f['Asy'], ry4['text']]
            for c, v in enumerate(vals, 1):
                cell = ws4.cell(row=i, column=c, value=v); cell.border = bd

        # 5. Sistemas de enlace
        ws5 = wb.create_sheet("Sistemas Enlace")
        ws5['A1'] = "SISTEMAS DE CIMENTACIÓN ENLAZADOS"; ws5['A1'].font = Font(bold=True, size=12)
        vh = ['Sistema', 'Dir', 'Zapatas', 'Nodos', 'L total(m)', 'Mu+(kNm)', 'Mu-(kNm)', 'Vu(kN)',
              'b(cm)', 'h(cm)', 'As inf(cm2)', 'Ref inf', 'As sup(cm2)', 'Ref sup', 'V%', 's_est(cm)', 'Estado']
        wh(ws5, 3, vh)
        vrow = 4
        for sys in results.get('tie_systems', []):
            if sys.get('status') in ('insuficiente',):
                continue
            vals = [sys['system_id'], sys['direction'], ', '.join(sys.get('footings', [])),
                    sys['num_nodes'], sys['total_length'],
                    sys['Mu_max_pos'], sys['Mu_max_neg'], sys['Vu_max'],
                    round(sys['b_viga']*100), round(sys['h_viga']*100),
                    sys['As_inf'], sys['As_inf_text'], sys['As_sup'], sys['As_sup_text'],
                    round(sys['sr_viga']*100, 1), round(sys['s_estribo']*100), sys['status']]
            for c, v in enumerate(vals, 1):
                cell = ws5.cell(row=vrow, column=c, value=v); cell.border = bd
                if c == len(vals):
                    cell.fill = okf if v == 'ok' else ff_fill
            vrow += 1

        buf = io.BytesIO(); wb.save(buf); buf.seek(0)
        st.download_button("📥 Informe Excel", data=buf.getvalue(),
            file_name="informe_cimentaciones.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", type="primary")
        st.success("✅ Informe generado")
